from __future__ import annotations

import csv
import os
import json
import math
import platform
import random
import socket
import struct
import subprocess
import sys
import threading
import time
import traceback
from dataclasses import dataclass, field
from pathlib import Path
from queue import Queue, Empty

from PyQt6.QtCore import Qt, QTimer, QPoint, QRectF, QPointF
from PyQt6.QtGui import QColor, QFont, QPainter, QPen, QBrush, QAction, QShortcut, QKeySequence, QIcon
from PyQt6.QtWidgets import (
    QApplication, QWidget, QMainWindow, QMenu, QVBoxLayout, QHBoxLayout, QGridLayout,
    QFormLayout, QLabel, QPushButton, QSpinBox, QDoubleSpinBox, QCheckBox, QComboBox,
    QColorDialog, QTabWidget, QLineEdit, QMessageBox, QGroupBox, QFrame, QFileDialog, QScrollArea
)
from PyQt6.QtWidgets import QScrollArea


try:
    import win32gui, win32con
    WIN32_OK = True
except Exception:
    WIN32_OK = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False


APP_NAME = "ONYX Drive HUD"
CONFIG_PATH = Path("onyx_drive_hud_config.json")
ICON_PATH = Path("onyx_icon.ico")
LOG_DIR = Path("logs")
def log_error(context, exc):
    pass

CRASH_LOG_PATH = LOG_DIR / "onyx_crash.log"

def log_error(where: str, exc: BaseException):
    try:
        LOG_DIR.mkdir(exist_ok=True)
        with CRASH_LOG_PATH.open("a", encoding="utf-8") as f:
            f.write("\n" + "=" * 80 + "\n")
            f.write(f"{time.strftime('%Y-%m-%d %H:%M:%S')} | {where}\n")
            f.write("".join(traceback.format_exception(type(exc), exc, exc.__traceback__)))
    except Exception:
        pass


LANGUAGES = {
    "en": "English",
    "de": "Deutsch",
    "es": "Español",
    "fr": "Français",
    "it": "Italiano",
    "pt": "Português",
    "pl": "Polski",
    "tr": "Türkçe",
    "ru": "Русский",
    "ja": "日本語",
    "ko": "한국어",
    "zh": "中文"
}

TEXT = {
    "en": {
        "title": "ONYX Drive HUD v5.2.7e Vehicle Database No-AI",
        "subtitle": "DRIVE HUD CONTROL CENTER · ONE PROCESS · PERFORMANCE LAB",
        "general": "General", "tiles": "Tiles", "peak": "Peak Measurements", "prototype": "Performance Lab",
        "design": "Design", "language": "Language", "hotkeys": "Keybinds", "units": "Units", "stability": "Stability",
        "save": "Save", "show_overlay": "Show Overlay", "hide_overlay": "Hide Overlay",
        "reset": "Reset all", "exit_app": "Exit ONYX", "start_recording": "Start Recording", "stop": "Stop",
        "reset_peaks": "Reset Peaks", "export_csv": "Export CSV", "export_xlsx": "Export XLSX",
        "system": "System / Appearance", "udp_port": "UDP port", "opacity": "Overlay opacity",
        "scale": "Scale", "tile_bg": "Tile background", "edit_mode": "Edit mode",
        "clickthrough": "Click-through", "select_tile": "Select tile",
        "position": "Position / Size / Color", "visible": "Visible", "width": "Width",
        "height": "Height", "label": "Label", "color": "Color", "choose_color": "Choose color",
        "apply_tile": "Apply tile", "theme": "Theme", "lang_hint": "English is default. German is included. Other major languages are available with mostly translated UI labels.",
        "saved": "Saved", "saved_msg": "Configuration saved. Overlay updates immediately.",
        "one_exe_note": "Overlay, Manager and Peak Measurements now share one UDP receiver in the same EXE.",
        "unit_system": "Unit system", "metric": "Metric", "imperial": "Imperial", "custom": "Custom",
        "speed_unit": "Speed unit", "power_unit": "Power unit", "boost_unit": "Boost unit", "gear_label": "Gear label",
        "units_hint": "Metric uses KMH, PS and bar. Imperial uses MPH, HP and PSI. Custom lets you choose each unit manually.",
        "crash_log": "Crash log", "crash_log_hint": "If ONYX catches an internal UI/Dyno error, details are written to logs/onyx_crash.log.",
        "boost_fix_note": "BoostFix: Forza raw boost is interpreted as PSI. PSI displays raw value; bar converts PSI / 14.5038.",
        "dyno_perf_note": "DynoPerformanceFix: graph and labels are throttled so long recordings do not freeze the manager window.",
    },
    "de": {
        "title": "ONYX Drive HUD v5.2.7e Vehicle Database No-AI",
        "subtitle": "DRIVE HUD CONTROL CENTER · ONE PROCESS · PERFORMANCE LAB",
        "general": "Allgemein", "tiles": "Kacheln", "peak": "Peak-Werte", "prototype": "Performance Lab",
        "design": "Design", "language": "Sprache", "hotkeys": "Tasten", "units": "Einheiten", "stability": "Stabilität",
        "save": "Speichern", "show_overlay": "Overlay anzeigen", "hide_overlay": "Overlay verstecken",
        "reset": "Alles zurücksetzen", "exit_app": "ONYX beenden", "start_recording": "Aufnahme starten", "stop": "Stop",
        "reset_peaks": "Peak-Werte zurücksetzen", "export_csv": "CSV exportieren", "export_xlsx": "XLSX exportieren",
        "system": "System / Darstellung", "udp_port": "UDP-Port", "opacity": "Overlay-Deckkraft",
        "scale": "Skalierung", "tile_bg": "Kachel-Hintergrund", "edit_mode": "Bearbeitungsmodus",
        "clickthrough": "Click-through", "select_tile": "Kachel auswählen",
        "position": "Position / Größe / Farbe", "visible": "Sichtbar", "width": "Breite",
        "height": "Höhe", "label": "Label", "color": "Farbe", "choose_color": "Farbe wählen",
        "apply_tile": "Kachel übernehmen", "theme": "Theme", "lang_hint": "Englisch ist Standard. Deutsch ist vollständig drin. Weitere Hauptsprachen sind mit den wichtigsten UI-Labels drin.",
        "saved": "Gespeichert", "saved_msg": "Konfiguration gespeichert. Overlay aktualisiert sich direkt.",
        "one_exe_note": "Overlay, Manager und Peak Measurements teilen sich jetzt einen UDP-Empfänger in derselben EXE.",
        "unit_system": "Einheitensystem", "metric": "Metrisch", "imperial": "Imperial", "custom": "Benutzerdefiniert",
        "speed_unit": "Geschwindigkeit", "power_unit": "Leistung", "boost_unit": "Ladedruck", "gear_label": "Gang-Label",
        "units_hint": "Metrisch nutzt KMH, PS und bar. Imperial nutzt MPH, HP und PSI. Benutzerdefiniert erlaubt freie Auswahl.",
        "crash_log": "Crash-Log", "crash_log_hint": "Wenn ONYX einen internen UI-/Dyno-Fehler abfängt, stehen Details in logs/onyx_crash.log.",
        "boost_fix_note": "BoostFix: Forza-Rohboost wird als PSI behandelt. PSI zeigt den Rohwert; bar rechnet PSI / 14.5038.",
        "dyno_perf_note": "DynoPerformanceFix: Graph und Labels werden gedrosselt, damit lange Aufnahmen das Manager-Fenster nicht einfrieren.",
    },
}

PARTIAL = {
    "es": {"language":"Idioma","save":"Guardar","show_overlay":"Mostrar overlay","hide_overlay":"Ocultar overlay","reset":"Restablecer","choose_color":"Elegir color"},
    "fr": {"language":"Langue","save":"Enregistrer","show_overlay":"Afficher overlay","hide_overlay":"Masquer overlay","reset":"Réinitialiser","choose_color":"Choisir couleur"},
    "it": {"language":"Lingua","save":"Salva","show_overlay":"Mostra overlay","hide_overlay":"Nascondi overlay","reset":"Ripristina","choose_color":"Scegli colore"},
    "pt": {"language":"Idioma","save":"Salvar","show_overlay":"Mostrar overlay","hide_overlay":"Ocultar overlay","reset":"Redefinir","choose_color":"Escolher cor"},
    "pl": {"language":"Język","save":"Zapisz","show_overlay":"Pokaż overlay","hide_overlay":"Ukryj overlay","reset":"Resetuj","choose_color":"Wybierz kolor"},
    "tr": {"language":"Dil","save":"Kaydet","show_overlay":"Overlay göster","hide_overlay":"Overlay gizle","reset":"Sıfırla","choose_color":"Renk seç"},
    "ru": {"language":"Язык","save":"Сохранить","show_overlay":"Показать overlay","hide_overlay":"Скрыть overlay","reset":"Сброс","choose_color":"Выбрать цвет"},
    "ja": {"language":"言語","save":"保存","show_overlay":"Overlay 表示","hide_overlay":"Overlay 非表示","reset":"リセット","choose_color":"色を選択"},
    "ko": {"language":"언어","save":"저장","show_overlay":"오버레이 표시","hide_overlay":"오버레이 숨김","reset":"초기화","choose_color":"색상 선택"},
    "zh": {"language":"语言","save":"保存","show_overlay":"显示 Overlay","hide_overlay":"隐藏 Overlay","reset":"重置","choose_color":"选择颜色"},
}

def tr(lang: str, key: str) -> str:
    if lang in TEXT and key in TEXT[lang]:
        return TEXT[lang][key]
    if lang in PARTIAL and key in PARTIAL[lang]:
        return PARTIAL[lang][key]
    return TEXT["en"].get(key, key)


THEMES = {
    "Blackout Blue": {"bg":"#05080d","panel":"#07111b","field":"#02060a","accent":"#00b8ff","accent2":"#00d9ff","text":"#d9f7ff"},
    "Purple Night": {"bg":"#080511","panel":"#120a22","field":"#05020a","accent":"#b84dff","accent2":"#d66bff","text":"#f1ddff"},
    "Emerald Grid": {"bg":"#03100b","panel":"#071b13","field":"#010806","accent":"#00ff99","accent2":"#55ffcc","text":"#ddfff3"},
    "Amber Track": {"bg":"#120a02","panel":"#1d1004","field":"#090500","accent":"#ff9d00","accent2":"#ffd15a","text":"#fff2dd"},
    "Glass Cyan": {"bg":"#03090d","panel":"#061620","field":"#010507","accent":"#53e8ff","accent2":"#9af4ff","text":"#ecfcff"},
}

DEFAULT_CONFIG = {
    "udp_host": "0.0.0.0",
    "udp_port": 5607,
    "language": "en",
    "manager_theme": "Blackout Blue",
    "unit_system": "Metric",
    "speed_unit": "KMH",
    "power_unit": "PS",
    "boost_unit": "bar",
    "gear_label": "GEAR",
    "dyno_ui_fps": 10,
    "dyno_max_samples": 20000,
    "dyno_zoom": 1.0,
    "dyno_record_mode": "full_record",
    "active_profile": "Default",
    "prototype_enabled": True,
    "performance_lab_enabled": True,
        "vehicle_badge_enabled": True,
    "vehicle_badge_position": "Top Right",
    "live_graph_paused": False,
    "drag_recording": False,
    "grip_recording": False,
    "hints_recording": False,
    "session_recording": False,
    "edit_mode": True,
    "click_through": False,
    "opacity": 0.92,
    "scale": 1.0,
    "background_alpha": 115,
    "overlay_monitor_index": -1,
    "font_family": "Segoe UI",
    "hotkeys": {"toggle_edit":"Ctrl+E","toggle_click":"Ctrl+T","save_layout":"Ctrl+S","reset_layout":"Ctrl+R","hide_overlay":"Esc","toggle_overlay":"F8","toggle_recording":"F9","reset_peaks":"F10","toggle_drag_record":"Ctrl+D","reset_drag":"Ctrl+Shift+D","toggle_grip_record":"Ctrl+G","analyze_grip":"Ctrl+Shift+G","reset_grip":"Alt+G","toggle_hints_record":"Ctrl+H","analyze_hints":"Ctrl+Shift+H","reset_hints":"Alt+H","toggle_session_record":"Ctrl+J","export_session":"Ctrl+Shift+J","reset_session":"Alt+J","pause_resume_live_graph":"Ctrl+L"},
    "cards": {
        "speed": {"x": 35, "y": 85, "w": 230, "h": 88, "label": "KMH", "color": "#00d9ff", "visible": True, "label_visible": True, "label_auto": True},
        "rpm": {"x": 35, "y": 180, "w": 230, "h": 88, "label": "RPM", "color": "#ff6a00", "visible": True, "label_visible": True, "label_auto": True},
        "gear": {"x": 35, "y": 275, "w": 230, "h": 88, "label": "GEAR", "color": "#55ff00", "visible": True, "label_visible": True, "label_auto": True},
        "power": {"x": 35, "y": 370, "w": 230, "h": 88, "label": "PS", "color": "#ffd400", "visible": True, "label_visible": True, "label_auto": True},
        "boost": {"x": 35, "y": 465, "w": 230, "h": 88, "label": "BOOST", "color": "#c74cff", "visible": True, "label_visible": True, "label_auto": True},
        "grip": {"x": 35, "y": 560, "w": 230, "h": 88, "label": "GRIP", "color": "#00ff99", "visible": True, "label_visible": True, "label_auto": True},
        "tachometer": {"x": 300, "y": 90, "w": 260, "h": 260, "label": "RPM GAUGE", "color": "#ff6a00", "visible": False, "label_visible": True, "label_auto": True},
        "drag_timer": {"x": 300, "y": 370, "w": 260, "h": 150, "label": "DRAG TIMER", "color": "#00d9ff", "visible": False, "label_visible": True, "label_auto": True},
    }
}

CARD_LABELS = {"speed":"Speed","rpm":"RPM","gear":"Gear","power":"Power","boost":"Boost","grip":"Grip Warning","tachometer":"RPM Gauge / Tachometer","drag_timer":"Drag Timer Overlay"}


def deep_copy(obj):
    return json.loads(json.dumps(obj))

def load_config():
    cfg = deep_copy(DEFAULT_CONFIG)
    if CONFIG_PATH.exists():
        try:
            user = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
            for k, v in user.items():
                if k not in ("cards","hotkeys"):
                    cfg[k] = v
            if "hotkeys" in user:
                cfg["hotkeys"].update(user["hotkeys"])
            if "cards" in user:
                for ck, cv in user["cards"].items():
                    if ck in cfg["cards"]:
                        cfg["cards"][ck].update(cv)
        except Exception:
            pass
    return cfg

def save_config(cfg):
    CONFIG_PATH.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")

def make_style(theme_name):
    t = THEMES.get(theme_name, THEMES["Blackout Blue"])
    return f"""
QWidget {{
    background-color: {t['bg']};
    color: {t['text']};
    font-family: Segoe UI;
    font-size: 10.5pt;
}}
QTabWidget::pane {{
    border: 1px solid {t['accent']};
    border-radius: 10px;
    background: {t['panel']};
}}
QTabBar::tab {{
    background: {t['panel']};
    color: {t['text']};
    border: 1px solid {t['accent']};
    padding: 9px 16px;
    margin-right: 4px;
    border-top-left-radius: 8px;
    border-top-right-radius: 8px;
}}
QTabBar::tab:selected {{
    background: {t['accent']};
    color: #001018;
}}
QGroupBox {{
    border: 1px solid {t['accent']};
    border-radius: 12px;
    margin-top: 14px;
    padding: 12px;
    background-color: {t['panel']};
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    left: 14px;
    padding: 0 6px;
    color: {t['accent2']};
    font-weight: bold;
}}
QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox {{
    background-color: {t['field']};
    border: 1px solid {t['accent']};
    border-radius: 7px;
    padding: 6px;
    color: #ffffff;
}}
QPushButton {{
    background-color: {t['panel']};
    color: {t['text']};
    border: 1px solid {t['accent']};
    border-radius: 9px;
    padding: 9px 14px;
    font-weight: bold;
}}
QPushButton:hover {{
    background-color: {t['field']};
    border: 1px solid {t['accent2']};
}}
QPushButton:pressed {{
    background-color: {t['accent']};
    color: #001018;
}}
QLabel#TitleLabel {{
    color: #ffffff;
    font-size: 24pt;
    font-weight: 800;
    letter-spacing: 3px;
}}
QLabel#SubtitleLabel {{
    color: {t['accent2']};
    font-size: 10pt;
}}
QFrame#Header {{
    background-color: {t['field']};
    border: 1px solid {t['accent']};
    border-radius: 16px;
}}
"""



def speed_value(tel, cfg):
    if tel is None:
        return 0.0
    kmh_attr = getattr(tel, "speed_kmh", 0.0)
    try:
        kmh = float(kmh_attr() if callable(kmh_attr) else kmh_attr)
    except Exception:
        kmh = 0.0
    if cfg.get("speed_unit", "KMH") == "MPH":
        return kmh * 0.621371
    return kmh

def speed_label(cfg):
    return "MPH" if cfg.get("speed_unit", "KMH") == "MPH" else "KMH"

def power_value(tel, cfg):
    if tel is None:
        return 0.0
    try:
        if hasattr(tel, "power_w"):
            ps = max(0.0, float(getattr(tel, "power_w", 0.0) or 0.0)) / 735.49875
        else:
            ps = max(0.0, float(getattr(tel, "power_ps", 0.0) or 0.0))
    except Exception:
        ps = 0.0
    unit = cfg.get("power_unit", "PS")
    if unit == "HP":
        return ps * 0.986320
    if unit == "kW":
        return ps * 0.735499
    return ps

def power_label(cfg):
    return cfg.get("power_unit", "PS")

def boost_value(tel, cfg):
    """
    HOTFIX v4.5:
    Forza Data Out / Dash boost is treated as PSI.
    Correct display:
    - PSI mode: raw boost value
    - bar mode: raw PSI / 14.5038

    Example: -11.02 PSI -> -0.76 bar.
    """
    if tel is None:
        return 0.0
    try:
        raw_psi = float(getattr(tel, "boost", 0.0) or 0.0)
    except Exception:
        raw_psi = 0.0
    if cfg.get("boost_unit", "bar") == "PSI":
        return raw_psi
    return raw_psi / 14.5038

def boost_label(cfg):
    return cfg.get("boost_unit", "bar")

@dataclass
class Telemetry:
    timestamp: float = 0.0
    raw_packet_size: int = 0
    car_ordinal: int | None = None
    car_class: int | None = None
    car_performance_index: int | None = None
    drivetrain_type: int | None = None
    num_cylinders: int | None = None
    car_group: int | None = None
    official_car_ordinal_raw: int | None = None
    vehicle_id_source: str = ""
    vehicle_id_confidence: str = ""
    vehicle_tuple_candidates: list = field(default_factory=list)
    car_id_probe_candidates: list = field(default_factory=list)
    engine_max_rpm: float = 0.0
    current_engine_rpm: float = 0.0
    velocity_x: float = 0.0
    velocity_y: float = 0.0
    velocity_z: float = 0.0
    speed_mps: float = 0.0
    power_w: float = 0.0
    torque_nm: float = 0.0
    boost: float = 0.0
    gear: int = 0
    accel: int = 0
    brake: int = 0
    steer: int = 0
    front_combined_slip: float = 0.0
    rear_combined_slip: float = 0.0
    front_slip_ratio: float = 0.0
    rear_slip_ratio: float = 0.0
    tire_slip_angle_front: float = 0.0
    tire_slip_angle_rear: float = 0.0

    @property
    def speed_kmh(self):
        if abs(self.speed_mps) > 0.01:
            return self.speed_mps * 3.6
        return math.sqrt(self.velocity_x**2 + self.velocity_y**2 + self.velocity_z**2) * 3.6

    @property
    def rpm(self):
        return int(round(self.current_engine_rpm))

    @property
    def power_ps(self):
        return int(round(max(0.0, self.power_w) / 735.49875))

    @property
    def throttle_pct(self):
        return max(0.0, min(100.0, self.accel / 255.0 * 100.0))

    @property
    def brake_pct(self):
        return max(0.0, min(100.0, self.brake / 255.0 * 100.0))


class ForzaParser:
    def parse(self, data: bytes):
        if len(data) < 200:
            return None
        t = Telemetry(timestamp=time.time(), raw_packet_size=len(data))
        o = 0
        def read(fmt):
            nonlocal o
            size = struct.calcsize("<" + fmt)
            if o + size > len(data):
                raise ValueError("short")
            val = struct.unpack_from("<" + fmt, data, o)[0]
            o += size
            return val
        try:
            read("i"); read("I")
            t.engine_max_rpm = read("f")
            read("f")
            t.current_engine_rpm = read("f")
            read("f"); read("f"); read("f")
            t.velocity_x = read("f"); t.velocity_y = read("f"); t.velocity_z = read("f")
            for _ in range(6): read("f")
            for _ in range(4): read("f")  # suspension normalized
            slip_ratio = [read("f") for _ in range(4)]
            t.front_slip_ratio = (abs(slip_ratio[0]) + abs(slip_ratio[1])) / 2
            t.rear_slip_ratio = (abs(slip_ratio[2]) + abs(slip_ratio[3])) / 2
            for _ in range(4): read("f")  # wheel rotation speed
            for _ in range(4): read("i")  # wheel on rumble strip
            for _ in range(4): read("f")  # wheel in puddle depth
            for _ in range(4): read("f")  # surface rumble
            slip_angle = [read("f") for _ in range(4)]
            t.tire_slip_angle_front = (abs(slip_angle[0]) + abs(slip_angle[1])) / 2
            t.tire_slip_angle_rear = (abs(slip_angle[2]) + abs(slip_angle[3])) / 2
            combined = [read("f") for _ in range(4)]
            t.front_combined_slip = (abs(combined[0]) + abs(combined[1])) / 2
            t.rear_combined_slip = (abs(combined[2]) + abs(combined[3])) / 2
            for _ in range(4): read("f")  # suspension travel meters
            if o + 20 <= len(data):
                # FH6 official Data Out layout: CarOrdinal begins after the Sled block.
                # Important: a raw 0 is not a usable vehicle ID; keep it as diagnostic
                # raw value but do not claim that vehicle ID 0 is real.
                raw_car_ordinal = read("i")
                t.official_car_ordinal_raw = raw_car_ordinal
                t.car_class = read("i")
                t.car_performance_index = read("i")
                t.drivetrain_type = read("i")
                t.num_cylinders = read("i")
                if raw_car_ordinal and raw_car_ordinal > 0:
                    t.car_ordinal = raw_car_ordinal
                    t.vehicle_id_source = "FH6 official CarOrdinal @ offset 212"
                    t.vehicle_id_confidence = "high"

            # FH6 adds CarGroup / SmashableVelDiff / SmashableMass after NumCylinders.
            # The official packet is 324 bytes. Read CarGroup by absolute offset so older
            # configs/parsers do not shift the rest of the Dash values by accident.
            try:
                if len(data) >= 236:
                    t.car_group = struct.unpack_from("<I", data, 232)[0]
            except Exception:
                pass

            # Vehicle tuple resolver: if raw official CarOrdinal is 0, test nearby known
            # layouts without guessing a name. A candidate only wins if the following
            # Class/PI/Drivetrain/Cylinders fields look like real Forza values.
            try:
                tuple_candidates = []
                for off in range(196, min(len(data) - 20, 280) + 1, 4):
                    try:
                        cid, cls, pi, drive, cyl = struct.unpack_from("<iiiii", data, off)
                        if 1 <= cid <= 5000000 and 0 <= cls <= 7 and 100 <= pi <= 999 and 0 <= drive <= 2 and 0 <= cyl <= 24:
                            item = {"offset": off, "car_id": cid, "class": cls, "pi": pi, "drive": drive, "cylinders": cyl}
                            tuple_candidates.append(item)
                    except Exception:
                        pass
                t.vehicle_tuple_candidates = tuple_candidates[:12]
                if (not t.car_ordinal) and tuple_candidates:
                    best = tuple_candidates[0]
                    t.car_ordinal = int(best["car_id"])
                    t.car_class = int(best["class"])
                    t.car_performance_index = int(best["pi"])
                    t.drivetrain_type = int(best["drive"])
                    t.num_cylinders = int(best["cylinders"])
                    t.vehicle_id_source = f"Vehicle tuple candidate @ offset {best['offset']}"
                    t.vehicle_id_confidence = "medium"
            except Exception:
                pass

            try:
                probe = []
                limit = min(len(data) - 4, 512)
                for off in range(0, limit + 1, 4):
                    val = struct.unpack_from("<i", data, off)[0]
                    if 1 <= val <= 5000000:
                        probe.append({"offset": off, "value": val})
                t.car_id_probe_candidates = probe[:80]
            except Exception:
                t.car_id_probe_candidates = []

            candidates = []
            for base in (244, 232, o):
                if base + 79 <= len(data):
                    try:
                        speed = struct.unpack_from("<f", data, base + 12)[0]
                        power = struct.unpack_from("<f", data, base + 16)[0]
                        torque = struct.unpack_from("<f", data, base + 20)[0]
                        boost = struct.unpack_from("<f", data, base + 40)[0] if base + 44 <= len(data) else 0.0
                        co = base + 68
                        accel = struct.unpack_from("<B", data, co + 3)[0]
                        brake = struct.unpack_from("<B", data, co + 4)[0]
                        gear = struct.unpack_from("<B", data, co + 7)[0]
                        steer = struct.unpack_from("<b", data, co + 8)[0]
                        score = 0
                        if 0 <= speed < 250: score += 3
                        if 0 <= gear <= 15: score += 1
                        if -127 <= steer <= 127: score += 1
                        candidates.append((score, speed, power, torque, boost, accel, brake, gear, steer))
                    except Exception:
                        pass
            if candidates:
                candidates.sort(reverse=True)
                _, t.speed_mps, t.power_w, t.torque_nm, t.boost, t.accel, t.brake, t.gear, t.steer = candidates[0]
            return t
        except Exception:
            return None


class UdpReceiver:
    def __init__(self, host, port, queue):
        self.host, self.port, self.queue = host, int(port), queue
        self.parser = ForzaParser()
        self.stop_event = threading.Event()
        self.thread = None
        self.sock = None
        self.raw_count = 0
        self.parsed_count = 0
        self.last_sender = "-"
        self.last_error = ""

    def start(self):
        self.stop()
        self.stop_event.clear()
        self.thread = threading.Thread(target=self._run, daemon=True)
        self.thread.start()

    def stop(self):
        self.stop_event.set()
        if self.sock:
            try: self.sock.close()
            except Exception: pass
        self.sock = None

    def _run(self):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            s.bind((self.host, self.port))
            s.settimeout(0.5)
            self.sock = s
            self.last_error = ""
            while not self.stop_event.is_set():
                try:
                    data, addr = s.recvfrom(4096)
                    self.raw_count += 1
                    self.last_sender = f"{addr[0]}:{addr[1]}"
                    tel = self.parser.parse(data)
                    if tel:
                        self.parsed_count += 1
                        try:
                            self.queue.put_nowait(tel)
                        except Exception:
                            try:
                                self.queue.get_nowait()
                            except Exception:
                                pass
                            try:
                                self.queue.put_nowait(tel)
                            except Exception:
                                pass
                except socket.timeout:
                    pass
                except OSError:
                    break
                except Exception as e:
                    self.last_error = str(e)
        except Exception as e:
            self.last_error = str(e)


def grip_warning_state(tel):
    """
    Returns: dict with text, percent, color hex, severity 0-4, blink bool.
    Uses Forza UDP slip values when available. Falls back to speed/steer/throttle heuristic.
    """
    try:
        if tel is None:
            return {"text": "NO DATA", "percent": 0, "color": "#00d9ff", "severity": 0, "blink": False}

        front = float(getattr(tel, "front_combined_slip", 0.0) or 0.0)
        rear = float(getattr(tel, "rear_combined_slip", 0.0) or 0.0)
        fr = float(getattr(tel, "front_slip_ratio", 0.0) or 0.0)
        rr = float(getattr(tel, "rear_slip_ratio", 0.0) or 0.0)
        speed = float(getattr(tel, "speed_kmh", 0.0) or 0.0)
        throttle = float(getattr(tel, "throttle_pct", 0.0) or 0.0)
        steer = abs(float(getattr(tel, "steer", 0.0) or 0.0))

        # Combined severity. Speed makes high-speed slip more critical.
        slip_front = max(front, fr * 2.1)
        slip_rear = max(rear, rr * 2.1)
        slip = max(slip_front, slip_rear)

        speed_factor = 1.0
        if speed > 250:
            speed_factor = 1.35
        elif speed > 180:
            speed_factor = 1.20
        elif speed < 40:
            speed_factor = 0.70

        effective = slip * speed_factor

        text = "GRIP OK"
        severity = 0

        if effective >= 2.35:
            severity = 4
            if slip_rear > slip_front * 1.12:
                text = "REAR SLIP"
            elif slip_front > slip_rear * 1.12:
                text = "FRONT SLIP"
            else:
                text = "LOW GRIP"
        elif effective >= 1.65:
            severity = 3
            if slip_rear > slip_front * 1.15:
                text = "OVERSTEER"
            elif slip_front > slip_rear * 1.15:
                text = "UNDERSTEER"
            else:
                text = "GRIP RISK"
        elif effective >= 1.05:
            severity = 2
            if slip_rear > slip_front * 1.20:
                text = "REAR WARN"
            elif slip_front > slip_rear * 1.20:
                text = "FRONT WARN"
            else:
                text = "GRIP WARN"
        elif speed > 180 and throttle > 80 and steer > 70:
            severity = 2
            text = "HIGH LOAD"

        # fallback if slip telemetry is zero/unavailable
        if slip == 0 and speed > 160 and throttle > 80 and steer > 80:
            severity = max(severity, 2)
            text = "HIGH LOAD"

        # percent is "remaining grip feeling", not physical tire model truth.
        percent = max(0, min(100, int(round(100 - min(100, effective * 38)))))

        if severity >= 4:
            color = "#ff1e1e"
        elif severity == 3:
            color = "#ff7a00"
        elif severity == 2:
            color = "#ffd400"
        else:
            color = "#00ff99"

        return {"text": text, "percent": percent, "color": color, "severity": severity, "blink": severity >= 4}
    except Exception as exc:
        log_error("grip_warning_state", exc)
        return {"text": "GRIP ERR", "percent": 0, "color": "#ff1e1e", "severity": 4, "blink": True}


def get_overlay_screen(config):
    """
    Returns the selected screen for the overlay.
    -1 = primary monitor
     0+ = QApplication.screens()[index]
    Falls back to primary if the saved monitor index is invalid.
    """
    try:
        screens = QApplication.screens()
        saved = int(config.get("overlay_monitor_index", -1))
        if saved >= 0 and saved < len(screens):
            return screens[saved]
        return QApplication.primaryScreen() or (screens[0] if screens else None)
    except Exception as exc:
        log_error("get_overlay_screen", exc)
        try:
            return QApplication.primaryScreen()
        except Exception:
            return None

def overlay_screen_label(index, screen):
    try:
        geo = screen.geometry()
        name = screen.name() if hasattr(screen, "name") else ""
        prefix = "Primary Monitor" if index == -1 else f"Monitor {index + 1}"
        if name:
            return f"{prefix} · {name} · {geo.width()}x{geo.height()}"
        return f"{prefix} · {geo.width()}x{geo.height()}"
    except Exception:
        return "Monitor"


class Card:
    def __init__(self, key, cfg):
        self.key = key
        self.cfg = cfg
        self.drag_offset = QPoint(0,0)

    def _scale(self):
        try:
            return float(self.cfg.get("_global_config", {}).get("scale", 1.0))
        except Exception:
            return 1.0

    @property
    def rect(self):
        return QRectF(float(self.cfg["x"]), float(self.cfg["y"]), float(self.cfg["w"]), float(self.cfg["h"]))

    def draw_rect(self):
        s = self._scale()
        return QRectF(
            float(self.cfg["x"]) * s,
            float(self.cfg["y"]) * s,
            float(self.cfg["w"]) * s,
            float(self.cfg["h"]) * s
        )

    def contains(self, p):
        return self.draw_rect().contains(float(p.x()), float(p.y()))

    def value(self, tel):
        cfg = self.cfg.get("_global_config", {})
        if tel is None:
            return {"speed":"0","rpm":"0","gear":"N","power":"0","boost":"0.00","grip":"--","tachometer":"0","drag_timer":"DRAG"}.get(self.key,"-")
        try:
            if self.key == "speed":
                return f"{speed_value(tel, cfg):.0f}"
            if self.key == "rpm":
                return f"{getattr(tel, 'rpm', 0):,}".replace(",", ".")
            if self.key == "gear":
                gear = getattr(tel, "gear", 0)
                return "R/N" if gear == 0 else str(gear)
            if self.key == "power":
                return f"{power_value(tel, cfg):.0f}"
            if self.key == "boost":
                decimals = 1 if cfg.get("boost_unit", "bar") == "PSI" else 2
                return f"{boost_value(tel, cfg):.{decimals}f}".replace(".", ",")
            if self.key == "grip":
                state = grip_warning_state(tel)
                return f"{state['percent']}%"
            if self.key == "tachometer":
                return f"{getattr(tel, 'rpm', 0):.0f}"
        except Exception as exc:
            log_error("Card.value", exc)
            return "-"
        return "-"


class OverlayWindow(QWidget):
    def __init__(self, manager):
        super().__init__()
        self.manager = manager
        self.config = manager.config
        self.cards = {k: Card(k, {**self.config["cards"][k], "_global_config": self.config}) for k in self.config["cards"]}
        self.telemetry = None
        self.selected = None
        self.shortcuts = []
        self.setWindowTitle(APP_NAME + " Overlay")
        if ICON_PATH.exists():
            self.setWindowIcon(QIcon(str(ICON_PATH)))
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self.setMouseTracking(True)
        self.apply_window_flags()
        self.register_shortcuts()
        self._blink_timer = QTimer(self)
        self._blink_timer.timeout.connect(self.update)
        self._blink_timer.start(160)

    def sync_config(self):
        self.config = self.manager.config
        self.cards = {k: Card(k, {**self.config["cards"][k], "_global_config": self.config}) for k in self.config["cards"]}
        self.apply_window_flags()
        self.register_shortcuts()
        self.update()

    def register_shortcuts(self):
        try:
            self.shortcuts = []
            hotkeys = self.manager.config.get("hotkeys", {})
            actions = {
                "toggle_edit": getattr(self.manager, "toggle_edit", None),
                "toggle_click": getattr(self.manager, "toggle_click", None),
                "save_layout": getattr(self.manager, "save_layout", None),
                "reset_layout": getattr(self.manager, "reset_layout", None),
                "hide_overlay": self.hide,
                "toggle_overlay": getattr(self.manager, "toggle_overlay_visibility", None),
                "toggle_recording": getattr(self.manager, "toggle_recording", None),
                "reset_peaks": getattr(self.manager, "reset_peak_recording", None),
                "pause_resume_live_graph": getattr(self.manager, "pause_resume_live_graph", None),
                "pause_resume_live_graph": getattr(self.manager, "pause_resume_live_graph", None),
            }
            for name, seq in hotkeys.items():
                if not seq:
                    continue
                action = actions.get(name)
                if action is None:
                    continue
                try:
                    sc = QShortcut(QKeySequence(seq), self)
                    sc.activated.connect(action)
                    self.shortcuts.append(sc)
                except Exception as exc:
                    log_error(f"OverlayWindow.register_shortcuts.{name}", exc)
        except Exception as exc:
            log_error("OverlayWindow.register_shortcuts", exc)


    def apply_window_flags(self):
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint | Qt.WindowType.Tool)
        self.setWindowOpacity(float(self.config["opacity"]))
        self.show()
        if WIN32_OK:
            hwnd = int(self.winId())
            ex = win32gui.GetWindowLong(hwnd, win32con.GWL_EXSTYLE) | win32con.WS_EX_LAYERED
            if self.config.get("click_through"):
                ex |= win32con.WS_EX_TRANSPARENT
            else:
                ex &= ~win32con.WS_EX_TRANSPARENT
            win32gui.SetWindowLong(hwnd, win32con.GWL_EXSTYLE, ex)

    def contextMenuEvent(self, event):
        menu = QMenu(self)
        for text, func in [
            ("Show Manager", self.manager.show),
            ("Reload Config", self.sync_config),
            ("Toggle edit mode", self.manager.toggle_edit),
            ("Toggle click-through", self.manager.toggle_click),
            ("Save layout", self.manager.save_config_now),
            ("Exit", QApplication.quit)
        ]:
            a = QAction(text, self)
            a.triggered.connect(func)
            menu.addAction(a)
        menu.exec(event.globalPos())

    def mousePressEvent(self, event):
        if not self.config.get("edit_mode", True) or event.button() != Qt.MouseButton.LeftButton:
            return
        pos = event.position().toPoint()
        for c in reversed(list(self.cards.values())):
            if c.cfg.get("visible", True) and c.contains(pos):
                self.selected = c
                r = c.draw_rect()
                c.drag_offset = QPoint(int(pos.x()-r.x()), int(pos.y()-r.y()))
                return

    def mouseMoveEvent(self, event):
        if not self.config.get("edit_mode", True) or not self.selected:
            return
        try:
            pos = event.position().toPoint()
            scale = max(0.1, float(self.config.get("scale", 1.0)))
            new_x = int((pos.x() - self.selected.drag_offset.x()) / scale)
            new_y = int((pos.y() - self.selected.drag_offset.y()) / scale)

            # Important:
            # Card.cfg is a draw-time copy. The real persistent data lives in
            # manager.config["cards"][key]. Update both, otherwise Save Layout
            # writes the old coordinates and tiles jump back after restart.
            self.selected.cfg["x"] = new_x
            self.selected.cfg["y"] = new_y

            key = self.selected.key
            if key in self.manager.config.get("cards", {}):
                self.manager.config["cards"][key]["x"] = new_x
                self.manager.config["cards"][key]["y"] = new_y

            self.config = self.manager.config
            self.update()
        except Exception as exc:
            log_error("OverlayWindow.mouseMoveEvent", exc)

    def mouseReleaseEvent(self, event):
        if self.selected:
            try:
                key = self.selected.key
                if key in self.manager.config.get("cards", {}):
                    self.manager.config["cards"][key]["x"] = int(self.selected.cfg.get("x", self.manager.config["cards"][key].get("x", 0)))
                    self.manager.config["cards"][key]["y"] = int(self.selected.cfg.get("y", self.manager.config["cards"][key].get("y", 0)))
                save_config(self.manager.config)
                self.config = self.manager.config
                self.manager.reload_forms_from_config()
            except Exception as exc:
                log_error("OverlayWindow.mouseReleaseEvent", exc)
        self.selected = None

    def paintEvent(self, event):
        try:
            p = QPainter(self)
            p.setRenderHint(QPainter.RenderHint.Antialiasing)
            if self.config.get("edit_mode", True):
                self.draw_hint(p)
            self.draw_vehicle_badge(p)
            for c in self.cards.values():
                if c.cfg.get("visible", True):
                    self.draw_card(p, c)
        except Exception as exc:
            log_error("OverlayWindow.paintEvent", exc)

    def draw_hint(self, p):
        p.setPen(QColor(0, 220, 255, 210))
        p.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        r = self.manager.receiver
        p.drawText(18, 28, "ONYX ONEEXE | EDIT MODE | one UDP receiver")
        p.setFont(QFont("Segoe UI", 8))
        p.drawText(18, 48, f"UDP {self.config['udp_port']} | Raw {r.raw_count} | Parsed {r.parsed_count} | {r.last_sender}")
        p.drawText(18, 66, vehicle_summary_text(self.telemetry, compact=True))

    def draw_vehicle_badge(self, p):
        try:
            if not bool(self.config.get("vehicle_badge_enabled", True)):
                return
            scale = float(self.config.get("scale", 1.0))
            w = max(360, int(430 * scale))
            h = max(58, int(66 * scale))
            margin = max(14, int(20 * scale))
            pos = str(self.config.get("vehicle_badge_position", "Top Right"))
            if "Bottom" in pos:
                y = self.height() - h - margin
            else:
                y = margin
            if "Left" in pos:
                x = margin
            else:
                x = self.width() - w - margin
            rect = QRectF(x, y, w, h)
            col = QColor(0, 217, 255)
            p.setPen(QPen(QColor(0, 217, 255, 150), max(1, int(1.4 * scale))))
            p.setBrush(QBrush(QColor(2, 8, 14, min(185, int(self.config.get("background_alpha", 115)) + 45))))
            p.drawRoundedRect(rect, max(10, int(14 * scale)), max(10, int(14 * scale)))
            p.setFont(QFont(self.config.get("font_family", "Segoe UI"), max(8, int(10 * scale)), QFont.Weight.Bold))
            p.setPen(QColor(0, 217, 255, 230))
            p.drawText(QRectF(x + 12*scale, y + 6*scale, w - 24*scale, 18*scale), Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, "CURRENT VEHICLE")
            p.setFont(QFont(self.config.get("font_family", "Segoe UI"), max(9, int(12 * scale)), QFont.Weight.Bold))
            p.setPen(QColor(235, 250, 255, 230))
            txt = vehicle_summary_text(self.telemetry, compact=True)
            if txt.startswith("Vehicle: "):
                txt = txt[len("Vehicle: "):]
            p.drawText(QRectF(x + 12*scale, y + 26*scale, w - 24*scale, h - 30*scale), Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, txt)
        except Exception as exc:
            log_error("OverlayWindow.draw_vehicle_badge", exc)

    def draw_drag_timer_card(self, p, c):
        try:
            scale = float(self.config.get("scale", 1.0))
            rect = c.draw_rect()
            x, y, w, h = rect.x(), rect.y(), rect.width(), rect.height()
            col = QColor(c.cfg.get("color", "#00d9ff"))
            bg_alpha = int(self.config.get("background_alpha", 115))

            # Very light glass panel, as requested: mirror of existing Drag Timer.
            radius = max(8, int(14 * scale))
            p.setPen(QPen(QColor(col.red(), col.green(), col.blue(), 115), max(1, int(1.4 * scale))))
            p.setBrush(QBrush(QColor(2, 8, 14, max(35, min(135, bg_alpha)))))
            p.drawRoundedRect(rect, radius, radius)

            title = c.cfg.get("label", "DRAG TIMER") if not bool(c.cfg.get("label_auto", True)) else "DRAG TIMER"
            if bool(c.cfg.get("label_visible", True)):
                p.setFont(QFont(self.config.get("font_family", "Segoe UI"), max(8, int(h * 0.13)), QFont.Weight.Bold))
                p.setPen(QColor(col.red(), col.green(), col.blue(), 220))
                p.drawText(QRectF(x + 10*scale, y + 5*scale, w - 20*scale, h * 0.18), Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, title)

            live = getattr(self.manager, "drag_overlay_times", {}) or {}
            record = getattr(self.manager, "drag_overlay_record_times", {}) or {}
            recording = bool(getattr(self.manager, "drag_overlay_recording", False))

            # Prefer current live times. If there are no live times yet, show last record results.
            source = live if live else record
            status = "REC" if recording else ("LIVE" if live else "WAIT")

            rows = ["0-100", "0-200", "100-200", "200-300"]
            top = y + (h * 0.23 if bool(c.cfg.get("label_visible", True)) else h * 0.10)
            row_h = max(18, (h - (top - y) - 12*scale) / 4.0)

            p.setFont(QFont(self.config.get("font_family", "Segoe UI"), max(8, int(row_h * 0.48)), QFont.Weight.Bold))
            for i, key in enumerate(rows):
                yy = top + i * row_h
                val = source.get(key)
                val_txt = f"{val:.2f}s" if isinstance(val, (int, float)) else "--"
                p.setPen(QColor(220, 245, 255, 210))
                p.drawText(QRectF(x + 12*scale, yy, w * 0.48, row_h), Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, key)
                p.setPen(col if isinstance(val, (int, float)) else QColor(160, 175, 185, 170))
                p.drawText(QRectF(x + w * 0.48, yy, w * 0.42, row_h), Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter, val_txt)

            p.setFont(QFont(self.config.get("font_family", "Segoe UI"), max(6, int(h * 0.075)), QFont.Weight.Bold))
            p.setPen(QColor(col.red(), col.green(), col.blue(), 150))
            p.drawText(QRectF(x + 10*scale, y + h - 20*scale, w - 20*scale, 16*scale), Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter, status)

            if self.config.get("edit_mode", True):
                handle = max(10, int(20 * scale))
                p.setPen(QPen(QColor(col.red(), col.green(), col.blue(), 75), max(1, int(2 * scale))))
                p.setBrush(Qt.BrushStyle.NoBrush)
                p.drawEllipse(QRectF(x + w - handle - 8 * scale, y + 8 * scale, handle, handle))
        except Exception as exc:
            log_error("OverlayWindow.draw_drag_timer_card", exc)

    def draw_tachometer_card(self, p, c):
        try:
            scale = float(self.config.get("scale", 1.0))
            rect = c.draw_rect()
            x, y, w, h = rect.x(), rect.y(), rect.width(), rect.height()
            size = min(w, h)
            cx = x + w / 2
            cy = y + h / 2
            radius = size * 0.42

            col = QColor(c.cfg.get("color", "#ff6a00"))
            bg_alpha = int(self.config.get("background_alpha", 115))

            rpm = 0.0
            max_rpm = 9000.0
            gear = "N"
            if self.telemetry is not None:
                try:
                    rpm = float(getattr(self.telemetry, "rpm", 0.0) or 0.0)
                    max_rpm = float(getattr(self.telemetry, "max_rpm", 0.0) or 0.0)
                    if max_rpm <= 1000:
                        max_rpm = 9000.0
                    g = getattr(self.telemetry, "gear", 0)
                    gear = "R/N" if g == 0 else str(g)
                except Exception:
                    pass

            pct = max(0.0, min(1.0, rpm / max_rpm))
            start_angle = 225.0
            sweep = 270.0
            needle_angle = start_angle - sweep * pct

            # outer glass panel
            p.setPen(Qt.PenStyle.NoPen)
            p.setBrush(QBrush(QColor(0, 0, 0, min(210, bg_alpha + 35))))
            p.drawRoundedRect(rect, max(12, int(18 * scale)), max(12, int(18 * scale)))

            # dial background
            p.setBrush(QBrush(QColor(2, 8, 14, min(230, bg_alpha + 25))))
            p.setPen(QPen(QColor(col.red(), col.green(), col.blue(), 130), max(1, int(2 * scale))))
            p.drawEllipse(QPointF(cx, cy), radius, radius)

            # arc helper
            arc_rect = QRectF(cx - radius, cy - radius, radius * 2, radius * 2)
            p.setPen(QPen(QColor(70, 80, 90, 150), max(4, int(8 * scale))))
            p.drawArc(arc_rect, int((start_angle - sweep) * 16), int(sweep * 16))

            p.setPen(QPen(col, max(4, int(8 * scale))))
            p.drawArc(arc_rect, int((start_angle - sweep * pct) * 16), int((sweep * pct) * 16))

            # redline last 15%
            p.setPen(QPen(QColor(255, 30, 30, 190), max(3, int(5 * scale))))
            p.drawArc(arc_rect.adjusted(8*scale,8*scale,-8*scale,-8*scale), int((start_angle - sweep) * 16), int((sweep * 0.15) * 16))

            # ticks and numbers
            p.setFont(QFont(self.config.get("font_family", "Segoe UI"), max(6, int(size * 0.045)), QFont.Weight.Bold))
            for i in range(0, 11):
                t = i / 10.0
                ang = start_angle - sweep * t
                rad = math.radians(ang)
                outer = radius * 0.92
                inner = radius * (0.78 if i % 2 == 0 else 0.84)
                x1 = cx + math.cos(rad) * inner
                y1 = cy - math.sin(rad) * inner
                x2 = cx + math.cos(rad) * outer
                y2 = cy - math.sin(rad) * outer
                p.setPen(QPen(QColor(230, 250, 255, 180), max(1, int(2 * scale))))
                p.drawLine(QPointF(x1, y1), QPointF(x2, y2))
                if i % 2 == 0:
                    num = int((max_rpm / 1000.0) * t)
                    tx = cx + math.cos(rad) * radius * 0.62
                    ty = cy - math.sin(rad) * radius * 0.62
                    p.setPen(QColor(230, 250, 255, 190))
                    p.drawText(QRectF(tx - 18, ty - 9, 36, 18), Qt.AlignmentFlag.AlignCenter, str(num))

            # needle
            rad = math.radians(needle_angle)
            nx = cx + math.cos(rad) * radius * 0.72
            ny = cy - math.sin(rad) * radius * 0.72
            p.setPen(QPen(QColor(255, 255, 255, 230), max(2, int(3 * scale))))
            p.drawLine(QPointF(cx, cy), QPointF(nx, ny))
            p.setBrush(QBrush(col))
            p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(QPointF(cx, cy), max(5, int(7 * scale)), max(5, int(7 * scale)))

            # central values
            p.setFont(QFont(self.config.get("font_family", "Segoe UI"), max(14, int(size * 0.13)), QFont.Weight.Bold))
            p.setPen(col)
            p.drawText(QRectF(x, cy - size * 0.10, w, size * 0.16), Qt.AlignmentFlag.AlignCenter, f"{rpm:.0f}")
            p.setFont(QFont(self.config.get("font_family", "Segoe UI"), max(7, int(size * 0.055)), QFont.Weight.Bold))
            p.setPen(QColor(230, 250, 255, 210))
            p.drawText(QRectF(x, cy + size * 0.04, w, size * 0.08), Qt.AlignmentFlag.AlignCenter, "RPM")

            p.setFont(QFont(self.config.get("font_family", "Segoe UI"), max(9, int(size * 0.08)), QFont.Weight.Bold))
            p.setPen(QColor(85, 255, 0, 220))
            p.drawText(QRectF(x, cy + size * 0.13, w, size * 0.10), Qt.AlignmentFlag.AlignCenter, f"GEAR {gear}")

            if bool(c.cfg.get("label_visible", True)):
                label = c.cfg.get("label", "RPM GAUGE") if not bool(c.cfg.get("label_auto", True)) else "RPM GAUGE"
                p.setFont(QFont(self.config.get("font_family", "Segoe UI"), max(7, int(size * 0.05)), QFont.Weight.Bold))
                p.setPen(QColor(230, 250, 255, 180))
                p.drawText(QRectF(x, y + h - size * 0.13, w, size * 0.08), Qt.AlignmentFlag.AlignCenter, label)

            # subtle edit handle only in edit mode
            if self.config.get("edit_mode", True):
                handle = max(10, int(22 * scale))
                p.setPen(QPen(QColor(col.red(), col.green(), col.blue(), 90), max(1, int(2 * scale))))
                p.setBrush(Qt.BrushStyle.NoBrush)
                p.drawEllipse(QRectF(x + w - handle - 10 * scale, y + 10 * scale, handle, handle))
        except Exception as exc:
            log_error("OverlayWindow.draw_tachometer_card", exc)

    def draw_card(self, p, c):
        try:
            if c.key == "drag_timer":
                self.draw_drag_timer_card(p, c)
                return
            if c.key == "tachometer":
                self.draw_tachometer_card(p, c)
                return
            scale = float(self.config.get("scale", 1.0))
            rect = c.draw_rect()
            x, y, w, h = rect.x(), rect.y(), rect.width(), rect.height()

            col = QColor(c.cfg.get("color", "#ffffff"))
            grip_state = None
            if c.key == "grip":
                grip_state = grip_warning_state(self.telemetry)
                col = QColor(grip_state.get("color", c.cfg.get("color", "#00ff99")))
                if grip_state.get("blink", False) and int(time.time() * 7) % 2 == 0:
                    col = QColor(255, 255, 255)

            # Dynamic dimensions: fixes small HUD scale overlap.
            radius = max(8, int(16 * scale))
            shadow = max(2, int(6 * scale))
            left_bar_x = x + max(7, 14 * scale)
            left_bar_top = y + max(8, 18 * scale)
            left_bar_bottom = y + h - max(8, 18 * scale)
            left_pad = max(30, int(45 * scale))
            right_pad = max(12, int(22 * scale))

            p.setPen(Qt.PenStyle.NoPen)
            p.setBrush(QBrush(QColor(0, 0, 0, min(220, int(self.config.get("background_alpha", 115)) + 45))))
            p.drawRoundedRect(rect.adjusted(shadow, shadow, shadow, shadow), radius, radius)

            p.setPen(QPen(QColor(col.red(), col.green(), col.blue(), 230), max(1, int(2 * scale))))
            p.setBrush(QBrush(QColor(2, 8, 14, int(self.config.get("background_alpha", 115)))))
            p.drawRoundedRect(rect, radius, radius)

            p.setPen(QPen(col, max(2, int(4 * scale))))
            p.drawLine(int(left_bar_x), int(left_bar_top), int(left_bar_x), int(left_bar_bottom))

            # Font sizes are based on actual scaled tile height, not fixed constants.
            if c.key == "rpm":
                value_font = int(max(15, min(42, h * 0.36)))
            elif c.key == "grip":
                value_font = int(max(14, min(34, h * 0.34)))
            else:
                value_font = int(max(16, min(44, h * 0.40)))
            label_font = int(max(8, min(16, h * 0.17)))

            value_rect = QRectF(x + left_pad, y + max(2, 4 * scale), max(10, w - left_pad - right_pad), h * 0.58)

            p.setFont(QFont(self.config.get("font_family", "Segoe UI"), value_font, QFont.Weight.Bold))
            p.setPen(QColor(235, 250, 255) if c.key == "speed" else col)

            if c.key == "grip" and grip_state is not None:
                p.drawText(value_rect, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, c.value(self.telemetry))
                p.setFont(QFont(self.config.get("font_family", "Segoe UI"), max(8, int(label_font)), QFont.Weight.Bold))
                p.setPen(col)
                sub_rect = QRectF(x + left_pad, y + h * 0.45, max(10, w - left_pad - right_pad), h * 0.25)
                p.drawText(sub_rect, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, grip_state.get("text", "GRIP"))
            else:
                p.drawText(value_rect, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, c.value(self.telemetry))

            # Label handling:
            # - label_visible False = no label
            # - label_auto True = unit-aware default labels
            # - label_auto False = custom text from Tiles tab
            show_label = bool(c.cfg.get("label_visible", True))
            if show_label:
                cfg = c.cfg.get("_global_config", {})
                if bool(c.cfg.get("label_auto", True)):
                    if c.key == "speed":
                        label = speed_label(cfg)
                    elif c.key == "power":
                        label = power_label(cfg)
                    elif c.key == "boost":
                        label = boost_label(cfg)
                    elif c.key == "gear":
                        label = cfg.get("gear_label", "GEAR")
                    elif c.key == "grip":
                        label = "GRIP"
                    else:
                        label = c.cfg.get("label", c.key.upper())
                else:
                    label = str(c.cfg.get("label", "")).strip()

                if label:
                    p.setFont(QFont(self.config.get("font_family", "Segoe UI"), label_font, QFont.Weight.Bold))
                    p.setPen(col)
                    label_h = max(12, h * 0.24)
                    label_rect = QRectF(x + left_pad, y + h - label_h - max(2, 4 * scale), max(10, w - left_pad - right_pad), label_h)
                    p.drawText(label_rect, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, label)

            # Edit handle: only visible while edit mode is enabled, much less intrusive.
            if self.config.get("edit_mode", True):
                handle_size = max(10, int(24 * scale))
                p.setPen(QPen(QColor(col.red(), col.green(), col.blue(), 95), max(1, int(2 * scale))))
                p.setBrush(Qt.BrushStyle.NoBrush)
                p.drawEllipse(QRectF(x + w - handle_size - max(7, 10 * scale), y + max(7, 10 * scale), handle_size, handle_size))
        except Exception as exc:
            log_error("OverlayWindow.draw_card", exc)


class DynoGraph(QWidget):
    """
    Clean dyno graph:
    - does NOT connect raw telemetry in time order
    - filters to full-throttle pull samples
    - bins by RPM and keeps the strongest/cleanest value per RPM bucket
    - draws sorted RPM curves only, so steering/shifting/coasting no longer creates rectangles
    """
    def __init__(self):
        super().__init__()
        self.samples = []
        self.config = DEFAULT_CONFIG
        self.zoom = 1.0
        self.setMinimumHeight(330)

    def set_zoom(self, zoom):
        try:
            self.zoom = max(0.6, min(3.0, float(zoom)))
        except Exception:
            self.zoom = 1.0
        self.update()

    def set_config(self, config):
        self.config = config
        try:
            self.zoom = max(0.6, min(3.0, float(config.get("dyno_zoom", self.zoom))))
        except Exception:
            pass

    def set_samples(self, samples):
        try:
            cap = int(self.config.get("dyno_max_samples", 20000))
        except Exception:
            cap = 20000
        self.samples = samples[-cap:]
        self.update()

    def _build_dyno_points(self):
        if not self.samples:
            return []

        # Clean pull filter. This removes steering/coasting/shifting garbage.
        valid = [
            s for s in self.samples
            if s.rpm >= 1200
            and s.throttle_pct >= 85
            and s.brake_pct <= 3
            and s.gear > 0
            and s.power_ps >= 0
            and s.speed_kmh >= 10
        ]

        # Fallback: if the user recorded too little clean full-throttle data.
        if len(valid) < 8:
            valid = [
                s for s in self.samples
                if s.rpm >= 1200 and s.gear > 0 and s.speed_kmh >= 10 and s.power_ps >= 0
            ]

        if len(valid) < 4:
            return []

        # Bin by RPM. For each RPM bucket, keep the highest PS sample.
        # This behaves much more like a real dyno curve.
        bucket_size = 150
        buckets = {}
        for s in valid:
            bucket = int(round(s.rpm / bucket_size) * bucket_size)
            old = buckets.get(bucket)
            if old is None or s.power_ps > old.power_ps:
                buckets[bucket] = s

        points = [buckets[k] for k in sorted(buckets)]
        if len(points) < 4:
            return []

        # Light smoothing over PS/NM/Boost to avoid telemetry spikes.
        smoothed = []
        for i, s in enumerate(points):
            lo = max(0, i - 1)
            hi = min(len(points), i + 2)
            window = points[lo:hi]
            # Create a lightweight copy-like object via original sample mutation-free wrapper.
            class P: pass
            p = P()
            p.rpm = s.rpm
            p.speed_kmh = sum(x.speed_kmh for x in window) / len(window)
            p.power_ps = sum(x.power_ps for x in window) / len(window)
            p.power_w = p.power_ps * 735.49875
            p.torque_nm = sum(x.torque_nm for x in window) / len(window)
            p.boost = sum(x.boost for x in window) / len(window)
            smoothed.append(p)
        return smoothed

    def paintEvent(self, event):
        try:
            self._paint_safe(event)
        except Exception as exc:
            log_error("DynoGraph.paintEvent", exc)
            try:
                p = QPainter(self)
                p.setPen(QColor(0, 217, 255))
                p.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
                p.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, "Dyno graph error skipped. See logs/onyx_crash.log")
            except Exception:
                pass

    def _paint_safe(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)

        rect = self.rect().adjusted(14, 14, -14, -14)
        plot = rect.adjusted(82, 48, -42, -72)

        bg = QColor(3, 9, 15)
        grid = QColor(0, 180, 255, 65)
        cyan = QColor(0, 217, 255)
        amber = QColor(255, 190, 0)
        purple = QColor(190, 80, 255)
        green = QColor(85, 255, 0)

        p.setPen(QPen(QColor(0, 184, 255), 1))
        p.setBrush(QBrush(bg))
        p.drawRoundedRect(rect, 16, 16)

        p.setPen(QPen(grid, 1))
        for i in range(7):
            x = plot.left() + plot.width() * i / 6
            p.drawLine(int(x), plot.top(), int(x), plot.bottom())
        for i in range(5):
            y = plot.top() + plot.height() * i / 4
            p.drawLine(plot.left(), int(y), plot.right(), int(y))

        # Axis frame
        p.setPen(QPen(QColor(0, 184, 255, 150), 1))
        p.drawRect(plot)

        p.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        p.setPen(QColor(220, 250, 255))
        p.drawText(rect.left() + 18, rect.top() + 24, "PEAK MEASUREMENTS · DYNO CLEAN VIEW · FULL RECORD")

        points = self._build_dyno_points()
        if len(points) < 4:
            p.setFont(QFont("Segoe UI", 12))
            p.drawText(plot, Qt.AlignmentFlag.AlignCenter, "No clean pull yet. Use one gear, full throttle, no braking.")
            return

        raw_min_rpm = max(0, min(s.rpm for s in points) - 250)
        raw_max_rpm = max(s.rpm for s in points) + 250
        rpm_center = (raw_min_rpm + raw_max_rpm) / 2
        rpm_half = max(500, (raw_max_rpm - raw_min_rpm) / 2) * self.zoom
        min_rpm = max(0, rpm_center - rpm_half)
        max_rpm = rpm_center + rpm_half

        max_ps = max(1, max(power_value(s, self.config) for s in points))
        max_nm = max(1, max(abs(s.torque_nm) for s in points))
        max_boost = max(0.01, max(abs(boost_value(s, self.config)) for s in points))
        y_max = max(max_ps, max_nm) * self.zoom
        max_boost = max_boost * self.zoom

        # Axis labels: left = power/torque scale, bottom = RPM scale.
        axis_pen = QColor(200, 235, 245)
        p.setFont(QFont("Segoe UI", 8))
        p.setPen(axis_pen)

        for i in range(5):
            ratio = i / 4
            val = y_max * (1 - ratio)
            y = plot.top() + plot.height() * ratio
            p.drawText(rect.left() + 8, int(y) - 7, 64, 16, Qt.AlignmentFlag.AlignRight, f"{val:.0f}")

        for i in range(7):
            ratio = i / 6
            val = min_rpm + (max_rpm - min_rpm) * ratio
            x = plot.left() + plot.width() * ratio
            p.drawText(int(x) - 34, plot.bottom() + 8, 68, 16, Qt.AlignmentFlag.AlignCenter, f"{val:.0f}")

        p.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
        p.setPen(QColor(0, 217, 255))
        p.drawText(rect.left() + 8, plot.top() - 24, 74, 18, Qt.AlignmentFlag.AlignRight, f"{power_label(self.config)} / NM")
        p.setPen(QColor(220, 250, 255))
        p.drawText(plot.center().x() - 50, rect.bottom() - 36, 100, 18, Qt.AlignmentFlag.AlignCenter, "RPM")
        p.setPen(QColor(190, 80, 255))
        p.drawText(plot.right() - 210, plot.top() - 24, 210, 18, Qt.AlignmentFlag.AlignRight, f"BOOST scaled · max {max_boost:.2f} {boost_label(self.config)}")

        def x_for(rpm):
            return plot.left() + (rpm - min_rpm) / max(1, (max_rpm - min_rpm)) * plot.width()

        def y_for_power(v):
            return plot.bottom() - v / max(1, y_max) * plot.height()

        def y_for_boost(v):
            # Boost as separate scaled overlay, lower intensity.
            return plot.bottom() - abs(v) / max_boost * plot.height()

        def draw_curve(values, color, width=3, boost=False):
            p.setPen(QPen(color, width))
            prev = None
            for s, value in values:
                y = y_for_boost(value) if boost else y_for_power(value)
                pt = QPointF(x_for(s.rpm), y)
                if prev:
                    p.drawLine(prev, pt)
                prev = pt

        # PS curve
        draw_curve([(s, power_value(s, self.config)) for s in points], cyan, 3, False)

        # NM curve
        draw_curve([(s, abs(s.torque_nm)) for s in points], amber, 2, False)

        # Boost curve, scaled to graph height
        if max_boost > 0.05:
            draw_curve([(s, boost_value(s, self.config)) for s in points], purple, 2, True)

        # Labels
        p.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
        p.setPen(cyan)
        p.drawText(plot.left() + 8, plot.top() + 20, power_label(self.config))
        p.setPen(amber)
        p.drawText(plot.left() + 48, plot.top() + 20, "NM")
        p.setPen(purple)
        p.drawText(plot.left() + 92, plot.top() + 20, "BOOST")
        p.setPen(QColor(220, 250, 255))
        p.drawText(plot.left(), rect.bottom() - 18, f"RPM {min_rpm:.0f} → {max_rpm:.0f}   Zoom {self.zoom:.1f}x   Mode: Full Record")

        peak_ps = max(points, key=lambda s: s.power_ps)
        peak_nm = max(points, key=lambda s: abs(s.torque_nm))
        peak_speed = max(self.samples, key=lambda s: s.speed_kmh) if self.samples else None
        peak_boost = max(points, key=lambda s: abs(s.boost))

        p.setPen(green)
        p.drawText(
            plot.left() + 160,
            rect.bottom() - 20,
            f"Peak {power_value(peak_ps, self.config):.0f} {power_label(self.config)} @ {peak_ps.rpm:.0f} rpm · {abs(peak_nm.torque_nm):.0f} NM · {boost_value(peak_boost, self.config):.2f} {boost_label(self.config)} · {speed_value(peak_speed, self.config):.1f} {speed_label(self.config)}"
        )


class PeakTab(QWidget):
    def __init__(self, manager):
        super().__init__()
        self.manager = manager
        self.samples = []
        self.live_run_samples = []
        self.recording = False
        self._last_ui_update = 0.0
        self._last_graph_update = 0.0
        self._in_pull = False
        self.build_ui()

    def build_ui(self):
        root = QVBoxLayout(self)
        box = QGroupBox("Peak Measurements / Dyno Lab")
        grid = QGridLayout(box)
        root.addWidget(box)
        self.labels = {}
        names = ["status","vehicle","samples","peak_speed","peak_rpm","peak_ps","peak_nm","100_200","200_300"]
        defaults = ["Status: stopped","Vehicle: waiting for telemetry","Samples: 0","Peak Speed: 0 km/h","Peak RPM: 0","Peak PS: 0","Peak NM: 0","100–200: -","200–300: -"]
        for i,(n,dv) in enumerate(zip(names,defaults)):
            lab = QLabel(dv)
            lab.setStyleSheet("font-weight:bold; color:#d9f7ff;")
            self.labels[n] = lab
            grid.addWidget(lab, i//2, i%2)
        row = QHBoxLayout()
        root.addLayout(row)
        self.btn_start = QPushButton(tr(self.manager.lang(), "start_recording"))
        self.btn_start.clicked.connect(self.start)
        row.addWidget(self.btn_start)
        self.btn_stop = QPushButton(tr(self.manager.lang(), "stop"))
        self.btn_stop.clicked.connect(self.stop)
        row.addWidget(self.btn_stop)
        self.btn_reset = QPushButton(tr(self.manager.lang(), "reset_peaks"))
        self.btn_reset.clicked.connect(self.clear)
        row.addWidget(self.btn_reset)
        self.btn_csv = QPushButton(tr(self.manager.lang(), "export_csv"))
        self.btn_csv.clicked.connect(self.export_csv)
        row.addWidget(self.btn_csv)
        self.btn_xlsx = QPushButton(tr(self.manager.lang(), "export_xlsx"))
        self.btn_xlsx.clicked.connect(self.export_xlsx)
        row.addWidget(self.btn_xlsx)

        zoom_row = QHBoxLayout()
        root.addLayout(zoom_row)
        self.btn_zoom_out = QPushButton("Dyno Zoom Out")
        self.btn_zoom_out.clicked.connect(self.zoom_out)
        zoom_row.addWidget(self.btn_zoom_out)
        self.btn_zoom_in = QPushButton("Dyno Zoom In")
        self.btn_zoom_in.clicked.connect(self.zoom_in)
        zoom_row.addWidget(self.btn_zoom_in)
        self.btn_zoom_reset = QPushButton("Reset Dyno Zoom")
        self.btn_zoom_reset.clicked.connect(self.zoom_reset)
        zoom_row.addWidget(self.btn_zoom_reset)

        self.graph = DynoGraph()
        self.graph.set_config(self.manager.config)
        self.graph.set_zoom(self.manager.config.get("dyno_zoom", 1.0))
        root.addWidget(self.graph, 1)
        hint = QLabel("Tip: Full Record mode keeps the whole recording as graph data and filters to throttle/performance samples. Left axis shows power/torque scale, bottom axis shows RPM.")
        hint.setWordWrap(True)
        root.addWidget(hint)

    def set_dyno_zoom(self, zoom):
        try:
            zoom = max(0.6, min(3.0, float(zoom)))
            self.manager.config["dyno_zoom"] = zoom
            self.graph.set_zoom(zoom)
            if hasattr(self.manager, "save_config"):
                self.manager.save_config()
        except Exception as exc:
            log_error("PeakTab.set_dyno_zoom", exc)

    def zoom_out(self):
        self.set_dyno_zoom(float(self.manager.config.get("dyno_zoom", 1.0)) + 0.25)

    def zoom_in(self):
        self.set_dyno_zoom(float(self.manager.config.get("dyno_zoom", 1.0)) - 0.25)

    def zoom_reset(self):
        self.set_dyno_zoom(1.0)

    def start(self):
        self.samples = []
        self.live_run_samples = []
        self._last_ui_update = 0.0
        self._last_graph_update = 0.0
        self._in_pull = False
        self.recording = True
        self.graph.set_samples([])
        self.labels["status"].setText("Status: Recording")

    def stop(self):
        self.recording = False
        self.labels["status"].setText("Status: stopped")

    def clear(self):
        self.samples = []
        self.live_run_samples = []
        self._in_pull = False
        self.graph.set_samples([])
        self.update_labels()

    def add_sample(self, t):
        try:
            if not self.recording or t is None:
                return

            now = time.time()
            max_samples = int(self.manager.config.get("dyno_max_samples", 20000))

            self.samples.append(t)
            if len(self.samples) > max_samples:
                self.samples = self.samples[-max_samples:]

            throttle = float(getattr(t, "throttle_pct", 0.0) or 0.0)
            brake = float(getattr(t, "brake_pct", 0.0) or 0.0)
            rpm = float(getattr(t, "rpm", 0.0) or 0.0)
            speed = float(getattr(t, "speed_kmh", 0.0) or 0.0)
            gear = int(getattr(t, "gear", 0) or 0)

            clean_pull = throttle >= 85 and brake <= 3 and rpm >= 1200 and speed >= 5 and gear > 0

            if clean_pull:
                if not self._in_pull:
                    self.live_run_samples = []
                    self._in_pull = True
                self.live_run_samples.append(t)
                if len(self.live_run_samples) > 5000:
                    self.live_run_samples = self.live_run_samples[-5000:]
            else:
                if throttle < 20 or brake > 10 or speed < 3:
                    self._in_pull = False

            ui_interval = 1.0 / max(1, int(self.manager.config.get("dyno_ui_fps", 10)))

            if now - self._last_ui_update >= ui_interval:
                self._last_ui_update = now
                self.update_labels()

            if now - self._last_graph_update >= ui_interval:
                self._last_graph_update = now
                try:
                    if hasattr(self.graph, "set_config"):
                        self.graph.set_config(self.manager.config)
                    # Full Record mode: keep the entire recording as graph data.
                    # DynoGraph still filters to useful throttle/performance samples internally.
                    draw_samples = self.samples
                    self.graph.set_samples(draw_samples)
                except Exception as exc:
                    log_error("PeakTab.graph_update", exc)

        except Exception as exc:
            log_error("PeakTab.add_sample", exc)

    def calc_accel(self, low, high):
        best = None
        start = None
        for s in self.samples:
            if start is None and s.speed_kmh >= low and s.throttle_pct > 80:
                start = s
            if start is not None and s.speed_kmh >= high:
                dt = s.timestamp - start.timestamp
                if dt > 0:
                    best = dt if best is None else min(best, dt)
                start = None
        return "-" if best is None else f"{best:.2f} s"

    def update_labels(self):
        if not self.samples:
            self.labels["vehicle"].setText(vehicle_summary_text(getattr(self.manager, "latest", None)))
            self.labels["samples"].setText("Samples: 0")
            self.labels["peak_speed"].setText("Peak Speed: 0 km/h")
            self.labels["peak_rpm"].setText("Peak RPM: 0")
            self.labels["peak_ps"].setText("Peak PS: 0")
            self.labels["peak_nm"].setText("Peak NM: 0")
            self.labels["100_200"].setText("100–200: -")
            self.labels["200_300"].setText("200–300: -")
            return
        self.labels["vehicle"].setText(vehicle_summary_text(self.samples[-1]))
        peak_speed = max(self.samples, key=lambda s:s.speed_kmh)
        peak_rpm = max(self.samples, key=lambda s:s.rpm)
        peak_ps = max(self.samples, key=lambda s:s.power_ps)
        peak_nm = max(self.samples, key=lambda s:abs(s.torque_nm))
        self.labels["samples"].setText(f"Samples: {len(self.samples)}")
        self.labels["peak_speed"].setText(f"Peak Speed: {speed_value(peak_speed, self.manager.config):.1f} {speed_label(self.manager.config)}")
        self.labels["peak_rpm"].setText(f"Peak RPM: {peak_rpm.rpm:.0f}")
        self.labels["peak_ps"].setText(f"Peak {power_label(self.manager.config)}: {power_value(peak_ps, self.manager.config):.0f} @ {peak_ps.rpm:.0f} rpm")
        self.labels["peak_nm"].setText(f"Peak NM: {abs(peak_nm.torque_nm):.0f} @ {peak_nm.rpm:.0f} rpm")
        self.labels["100_200"].setText(f"100–200: {self.calc_accel(100,200)}")
        self.labels["200_300"].setText(f"200–300: {self.calc_accel(200,300)}")

    def rows(self):
        if not self.samples: return []
        t0 = self.samples[0].timestamp
        cfg = self.manager.config
        su = speed_label(cfg).lower()
        pu = power_label(cfg).lower()
        bu = boost_label(cfg).lower()
        rows = []
        for s in self.samples:
            try:
                rows.append({
                    "time_s": round((getattr(s, "timestamp", t0) or t0)-t0,4),
                    f"speed_{su}": round(speed_value(s, cfg),3),
                    "rpm": round(float(getattr(s, "rpm", 0.0) or 0.0),1),
                    f"power_{pu}": round(power_value(s, cfg),3),
                    "torque_nm": round(float(getattr(s, "torque_nm", 0.0) or 0.0),3),
                    "gear": getattr(s, "gear", 0),
                    "throttle_pct": round(float(getattr(s, "throttle_pct", 0.0) or 0.0),2),
                    "brake_pct": round(float(getattr(s, "brake_pct", 0.0) or 0.0),2),
                    f"boost_{bu}": round(boost_value(s, cfg),4),
                })
            except Exception as exc:
                log_error("PeakTab.rows.sample_skipped", exc)
        return rows

    def export_csv(self):
        rows = self.rows()
        if not rows:
            QMessageBox.warning(self, "No data", "No samples.")
            return
        path,_ = QFileDialog.getSaveFileName(self, "Export CSV", "onyx_peak_measurements.csv", "CSV (*.csv)")
        if not path: return
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()), delimiter=";")
            w.writeheader(); w.writerows(rows)
        QMessageBox.information(self, "Export", f"CSV saved:\n{path}")

    def export_xlsx(self):
        rows = self.rows()
        if not rows:
            QMessageBox.warning(self, "No data", "No samples.")
            return
        if not OPENPYXL_OK:
            QMessageBox.warning(self, "Missing dependency", "openpyxl missing.")
            return
        path,_ = QFileDialog.getSaveFileName(self, "Export XLSX", "onyx_peak_measurements.xlsx", "Excel (*.xlsx)")
        if not path: return
        wb = Workbook()
        ws = wb.active
        ws.title = "Peak Measurements"
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r[h] for h in headers])
        fill = PatternFill("solid", fgColor="00111A")
        font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="00B8FF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for c in ws[1]:
            c.fill = fill; c.font = font; c.border = border; c.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 22)
        summary = wb.create_sheet("Summary")
        self.labels["vehicle"].setText(vehicle_summary_text(self.samples[-1]))
        peak_speed = max(self.samples, key=lambda s:s.speed_kmh)
        peak_rpm = max(self.samples, key=lambda s:s.rpm)
        peak_ps = max(self.samples, key=lambda s:s.power_ps)
        peak_nm = max(self.samples, key=lambda s:abs(s.torque_nm))
        cfg = self.manager.config
        for r in [
            ["Metric","Value"],
            [f"Peak Speed {speed_label(cfg)}", round(speed_value(peak_speed, cfg),2)],
            ["Peak RPM",round(peak_rpm.rpm,0)],
            [f"Peak {power_label(cfg)}", round(power_value(peak_ps, cfg),2)],
            [f"Peak {power_label(cfg)} RPM",round(peak_ps.rpm,0)],
            ["Peak Torque NM",round(abs(peak_nm.torque_nm),2)],
            ["Peak Torque RPM",round(peak_nm.rpm,0)],
            ["100-200 km/h",self.calc_accel(100,200)],
            ["200-300 km/h",self.calc_accel(200,300)],
            ["Samples",len(self.samples)]
        ]:
            summary.append(r)
        wb.save(path)
        QMessageBox.information(self, "Export", f"XLSX saved:\n{path}")



class LiveGraphWidget(QWidget):
    """
    Time-based live telemetry graph with separate per-channel normalization.
    This fixes the old issue where RPM dominated the graph and speed/boost looked flat.
    """
    def __init__(self):
        super().__init__()
        self.samples = []
        self.config = DEFAULT_CONFIG
        self.setMinimumHeight(260)

    def set_config(self, config):
        self.config = config

    def set_samples(self, samples):
        self.samples = samples[-800:]
        self.update()

    def _series(self, samples):
        cfg = self.config
        return {
            "speed": [speed_value(s, cfg) for s in samples],
            "rpm": [float(getattr(s, "rpm", 0.0) or 0.0) for s in samples],
            "power": [power_value(s, cfg) for s in samples],
            "boost": [boost_value(s, cfg) for s in samples],
        }

    def _safe_max(self, vals, fallback=1.0):
        try:
            m = max(abs(float(v)) for v in vals) if vals else fallback
            return max(m, fallback)
        except Exception:
            return fallback

    def _nice_cap(self, key, vals):
        cfg = self.config
        raw = self._safe_max(vals, 1.0)

        if key == "speed":
            if cfg.get("speed_unit", "KMH") == "MPH":
                return max(120.0, min(300.0, raw * 1.12))
            return max(200.0, min(500.0, raw * 1.12))

        if key == "rpm":
            return max(8000.0, min(14000.0, raw * 1.05))

        if key == "power":
            unit = cfg.get("power_unit", "PS")
            base = 700.0 if unit in ("PS", "HP") else 500.0
            return max(base, min(2500.0, raw * 1.15))

        if key == "boost":
            unit = cfg.get("boost_unit", "bar")
            if unit == "PSI":
                return max(20.0, min(80.0, raw * 1.20))
            # boost can be negative at idle/vacuum, but positive boost should define top scale
            positives = [v for v in vals if v > 0]
            pos = self._safe_max(positives, 1.0)
            return max(1.0, min(5.0, pos * 1.25))

        return raw

    def _map_y(self, value, cap, y0, gh, key):
        # Boost supports vacuum/negative values: baseline is around 70% height.
        if key == "boost":
            # Clamp from -1.0 bar/~-15 PSI to positive cap.
            neg_min = -15.0 if self.config.get("boost_unit", "bar") == "PSI" else -1.0
            v = max(neg_min, min(float(value), cap))
            span = cap - neg_min if cap > neg_min else 1.0
            norm = (v - neg_min) / span
            return y0 + gh - norm * gh

        v = max(0.0, min(float(value), cap))
        return y0 + gh - (v / cap) * gh

    def paintEvent(self, event):
        try:
            p = QPainter(self)
            p.setRenderHint(QPainter.RenderHint.Antialiasing)

            w = self.width()
            h = self.height()
            margin_l = 58
            margin_r = 24
            margin_t = 38
            margin_b = 42
            x0 = margin_l
            y0 = margin_t
            gw = max(10, w - margin_l - margin_r)
            gh = max(10, h - margin_t - margin_b)

            p.fillRect(self.rect(), QColor(2, 8, 14, 180))

            # outer graph box
            p.setPen(QPen(QColor(0, 217, 255, 150), 1))
            p.setBrush(Qt.BrushStyle.NoBrush)
            p.drawRect(int(x0), int(y0), int(gw), int(gh))

            # grid
            p.setPen(QPen(QColor(0, 217, 255, 45), 1))
            for i in range(1, 5):
                yy = y0 + gh * i / 5
                p.drawLine(int(x0), int(yy), int(x0 + gw), int(yy))
            for i in range(1, 6):
                xx = x0 + gw * i / 6
                p.drawLine(int(xx), int(y0), int(xx), int(y0 + gh))

            p.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
            p.setPen(QColor(230, 250, 255))
            p.drawText(12, 22, "LIVE TELEMETRY GRAPH · TIME MODE · SEPARATE SCALES")

            if not self.samples:
                p.setFont(QFont("Segoe UI", 10))
                p.drawText(QRectF(x0, y0, gw, gh), Qt.AlignmentFlag.AlignCenter, "waiting for telemetry")
                return

            samples = self.samples
            series = self._series(samples)
            n = len(samples)

            colors = {
                "speed": QColor(0, 217, 255),
                "rpm": QColor(255, 106, 0),
                "power": QColor(255, 212, 0),
                "boost": QColor(199, 76, 255),
            }

            labels = {
                "speed": f"Speed ({speed_label(self.config)})",
                "rpm": "RPM",
                "power": f"Power ({power_label(self.config)})",
                "boost": f"Boost ({boost_label(self.config)})",
            }

            caps = {k: self._nice_cap(k, vals) for k, vals in series.items()}

            # legend with caps so the user knows each curve has its own scale.
            lx = x0 + 4
            for key in ["speed", "rpm", "power", "boost"]:
                p.setPen(colors[key])
                p.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
                cap_txt = f"{caps[key]:.0f}" if key != "boost" else f"{caps[key]:.1f}"
                p.drawText(int(lx), int(y0 - 8), f"{labels[key]} / max {cap_txt}")
                lx += 165

            # left axis is normalized percent because curves use separate scales.
            p.setPen(QColor(230, 250, 255))
            p.setFont(QFont("Segoe UI", 8))
            p.drawText(8, int(y0 + 4), "100%")
            p.drawText(20, int(y0 + gh / 2 + 4), "50%")
            p.drawText(31, int(y0 + gh + 4), "0%")
            p.save()
            p.translate(14, y0 + gh / 2 + 35)
            p.rotate(-90)
            p.drawText(0, 0, "separate scale per curve")
            p.restore()

            if n < 2:
                return

            def x_for(i):
                return x0 + (i / max(1, n - 1)) * gw

            # draw curves
            for key in ["speed", "rpm", "power", "boost"]:
                vals = series[key]
                cap = caps[key]
                p.setPen(QPen(colors[key], 2))
                last = None
                for i, v in enumerate(vals):
                    pt = QPointF(x_for(i), self._map_y(v, cap, y0, gh, key))
                    if last is not None:
                        p.drawLine(last, pt)
                    last = pt

            # time axis
            first_ts = getattr(samples[0], "timestamp", None)
            last_ts = getattr(samples[-1], "timestamp", None)
            duration = 0.0
            if first_ts is not None and last_ts is not None:
                try:
                    duration = max(0.0, float(last_ts) - float(first_ts))
                except Exception:
                    duration = 0.0

            p.setPen(QColor(230, 250, 255))
            p.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
            p.drawText(QRectF(x0, y0 + gh + 16, gw, 22), Qt.AlignmentFlag.AlignCenter, f"TIME · {duration:.1f}s window")

            # bottom tick labels
            p.setFont(QFont("Segoe UI", 8))
            for i in range(0, 7):
                xx = x0 + gw * i / 6
                sec = duration * i / 6
                p.setPen(QColor(0, 217, 255, 70))
                p.drawLine(int(xx), int(y0 + gh), int(xx), int(y0 + gh + 5))
                p.setPen(QColor(230, 250, 255, 180))
                p.drawText(QRectF(xx - 24, y0 + gh + 24, 48, 16), Qt.AlignmentFlag.AlignCenter, f"{sec:.0f}s")

            # latest values, right side
            last_sample = samples[-1]
            latest_lines = [
                f"{speed_label(self.config)} {speed_value(last_sample, self.config):.0f}",
                f"RPM {getattr(last_sample, 'rpm', 0):.0f}",
                f"{power_label(self.config)} {power_value(last_sample, self.config):.0f}",
                f"{boost_label(self.config)} {boost_value(last_sample, self.config):.2f}",
            ]
            p.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
            yy = y0 + 14
            for key, txt in zip(["speed", "rpm", "power", "boost"], latest_lines):
                p.setPen(colors[key])
                p.drawText(int(x0 + gw - 120), int(yy), txt)
                yy += 16

        except Exception as exc:
            log_error("LiveGraphWidget.paintEvent", exc)



VEHICLE_DB_PATH = Path("car_database.json")
VEHICLE_DB_GENERATED_PATH = Path("car_database_generated_cleaned.json")
VEHICLE_DB_CUSTOM_PATH = Path("car_database_custom.json")
UNKNOWN_VEHICLES_PATH = Path("unknown_vehicles.json")

_VEHICLE_DB_CACHE = None
_VEHICLE_DB_CACHE_MTIMES = None


def _load_vehicle_db_file(path: Path):
    try:
        if not path.exists():
            return {}
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict) and isinstance(data.get("cars"), dict):
            return data.get("cars", {})
        if isinstance(data, dict) and isinstance(data.get("vehicles"), dict):
            return data.get("vehicles", {})
        if isinstance(data, dict):
            return data
    except Exception as exc:
        log_error(f"load_vehicle_database:{path}", exc)
    return {}


def load_vehicle_database(force_reload: bool = False):
    """Load CarOrdinal -> vehicle metadata.

    Priority:
    1) car_database_generated_cleaned.json / car_database.json = generated game-asset scan
    2) car_database_custom.json = user/community corrections, overrides generated names

    This keeps ONYX free/offline and avoids AI/API/Docker requirements.
    """
    global _VEHICLE_DB_CACHE, _VEHICLE_DB_CACHE_MTIMES
    paths = [VEHICLE_DB_GENERATED_PATH, VEHICLE_DB_PATH, VEHICLE_DB_CUSTOM_PATH]
    mtimes = tuple((str(p), p.stat().st_mtime if p.exists() else None) for p in paths)
    if (not force_reload) and _VEHICLE_DB_CACHE is not None and _VEHICLE_DB_CACHE_MTIMES == mtimes:
        return _VEHICLE_DB_CACHE
    db = {}
    for path in paths:
        part = _load_vehicle_db_file(path)
        for k, v in part.items():
            try:
                # Keep only numeric car IDs as lookup entries.
                key = str(int(k))
            except Exception:
                continue
            db[key] = v
    _VEHICLE_DB_CACHE = db
    _VEHICLE_DB_CACHE_MTIMES = mtimes
    return db


def vehicle_entry_for_id(car_id):
    if car_id is None:
        return None
    db = load_vehicle_database()
    return db.get(str(car_id))


def vehicle_name_for_id(car_id):
    entry = vehicle_entry_for_id(car_id)
    if isinstance(entry, str):
        return entry
    if isinstance(entry, dict):
        for k in ("display_name", "name", "vehicle", "car"):
            if entry.get(k):
                return str(entry.get(k))
        year = entry.get("year")
        make = entry.get("make") or entry.get("manufacturer")
        model = entry.get("model") or entry.get("raw_model")
        parts = []
        if year:
            parts.append(str(year))
        if make:
            parts.append(str(make))
        if model:
            parts.append(str(model))
        if parts:
            return " ".join(parts)
        if entry.get("asset"):
            return str(entry.get("asset"))
    return None


def vehicle_asset_for_id(car_id):
    entry = vehicle_entry_for_id(car_id)
    if isinstance(entry, dict):
        return entry.get("asset") or entry.get("zip_file") or ""
    return ""


def vehicle_database_stats():
    try:
        db = load_vehicle_database()
        generated = len(_load_vehicle_db_file(VEHICLE_DB_GENERATED_PATH if VEHICLE_DB_GENERATED_PATH.exists() else VEHICLE_DB_PATH))
        custom = len(_load_vehicle_db_file(VEHICLE_DB_CUSTOM_PATH))
        return len(db), generated, custom
    except Exception as exc:
        log_error("vehicle_database_stats", exc)
        return 0, 0, 0


def vehicle_class_label(cls):
    try:
        mapping = {0: "D", 1: "C", 2: "B", 3: "A", 4: "S1", 5: "S2", 6: "X", 7: "X"}
        return mapping.get(int(cls), str(cls))
    except Exception:
        return "-"


def drivetrain_label(v):
    try:
        return {0: "FWD", 1: "RWD", 2: "AWD"}.get(int(v), str(v))
    except Exception:
        return "-"


def effective_car_id(tel):
    try:
        cid = getattr(tel, "car_ordinal", None)
        if cid is not None and int(cid) > 0:
            return int(cid)
    except Exception:
        pass
    return None


def vehicle_display_name(tel):
    cid = effective_car_id(tel)
    if cid is None:
        return None
    return vehicle_name_for_id(cid)


def vehicle_summary_text(tel, compact=False):
    if tel is None:
        return "Vehicle: waiting for telemetry"
    try:
        cid = effective_car_id(tel)
        raw = getattr(tel, "official_car_ordinal_raw", None)
        name = vehicle_display_name(tel)
        if cid is None:
            if raw == 0:
                return "Vehicle: unknown · Car ID unavailable · official raw 0"
            return "Vehicle: unknown · Car ID unavailable"
        cls = vehicle_class_label(getattr(tel, "car_class", None))
        pi = getattr(tel, "car_performance_index", None)
        drive = drivetrain_label(getattr(tel, "drivetrain_type", None))
        display = name or "Unknown vehicle"
        base = f"Vehicle: {display} · ID {cid}"
        if pi is not None:
            base += f" · {cls} {pi}"
        if not compact:
            base += f" · {drive}"
            cyl = getattr(tel, "num_cylinders", None)
            if cyl is not None:
                base += f" · {cyl} cyl"
            group = getattr(tel, "car_group", None)
            if group not in (None, 0):
                base += f" · Group {group}"
            asset = vehicle_asset_for_id(cid)
            if asset:
                base += f" · {asset}"
            src = getattr(tel, "vehicle_id_source", "")
            conf = getattr(tel, "vehicle_id_confidence", "")
            if src:
                base += f" · {conf or 'detected'}"
        return base
    except Exception as exc:
        log_error("vehicle_summary_text", exc)
        return "Vehicle: status error"


def current_vehicle_unknown_entry(tel):
    cid = effective_car_id(tel)
    if cid is None:
        return None
    if vehicle_name_for_id(cid):
        return None
    return {
        "car_id": int(cid),
        "display_name": "",
        "year": None,
        "make": "",
        "model": "",
        "asset": "",
        "class": vehicle_class_label(getattr(tel, "car_class", None)),
        "pi": getattr(tel, "car_performance_index", None),
        "drivetrain": drivetrain_label(getattr(tel, "drivetrain_type", None)),
        "cylinders": getattr(tel, "num_cylinders", None),
        "car_group": getattr(tel, "car_group", None),
        "source": "ONYX unknown vehicle logger",
        "confidence": "needs-user-confirmation",
        "notes": "Fill display_name/year/make/model, then copy this entry into car_database_custom.json."
    }


def log_unknown_vehicle(tel):
    entry = current_vehicle_unknown_entry(tel)
    if not entry:
        return False
    try:
        data = {"unknown_vehicles": {}}
        if UNKNOWN_VEHICLES_PATH.exists():
            existing = json.loads(UNKNOWN_VEHICLES_PATH.read_text(encoding="utf-8"))
            if isinstance(existing, dict):
                data.update(existing)
            if not isinstance(data.get("unknown_vehicles"), dict):
                data["unknown_vehicles"] = {}
        key = str(entry["car_id"])
        old = data["unknown_vehicles"].get(key, {})
        if isinstance(old, dict):
            old.update(entry)
            entry = old
        entry["last_seen"] = time.strftime("%Y-%m-%d %H:%M:%S")
        data["unknown_vehicles"][key] = entry
        UNKNOWN_VEHICLES_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return True
    except Exception as exc:
        log_error("log_unknown_vehicle", exc)
        return False


def unknown_vehicle_clipboard_text(tel):
    entry = current_vehicle_unknown_entry(tel)
    if not entry:
        cid = effective_car_id(tel)
        if cid is None:
            return "No active vehicle ID detected yet."
        name = vehicle_name_for_id(cid)
        return f"Vehicle already known: {cid} = {name}"
    return json.dumps({str(entry["car_id"]): entry}, ensure_ascii=False, indent=2)


def summarize_vehicle_id_probe(samples, max_items=8):
    """Return stable integer candidates seen in the current UDP packet stream.

    Official Forza CarOrdinal is the only value ONYX treats as a confirmed car ID.
    Probe candidates are diagnostics only: switch cars and compare whether one
    offset/value changes with the vehicle while normal telemetry stays valid.
    """
    if not samples:
        return []
    counts = {}
    window = samples[-2000:]
    window_total = len(window)
    for s in window:
        for item in getattr(s, "car_id_probe_candidates", []) or []:
            try:
                key = (int(item.get("offset")), int(item.get("value")))
            except Exception:
                continue
            counts[key] = counts.get(key, 0) + 1
    if not counts:
        return []
    result = []
    for (offset, value), count in counts.items():
        if count >= max(20, int(window_total * 0.75)):
            result.append({"offset": offset, "value": value, "seen": count})
    result.sort(key=lambda x: (-x["seen"], x["offset"], x["value"]))
    return result[:max_items]


class PrototypeLabTab(QWidget):
    """
    Performance Lab Prototype:
    - Live Telemetry Graph
    - Drag Timer: Live + Manual Record
    - Grip Monitor: Live + Manual Record
    - Smart Hints: Live + Manual Record
    - Session Report: Live + Manual Record
    - HUD Presets
    - Profiles
    - Support Info
    """
    def __init__(self, manager):
        super().__init__()
        self.manager = manager

        self.samples = []
        self.drag_samples = []
        self.grip_samples = []
        self.hints_samples = []
        self.session_samples = []

        self.drag_recording = False
        self.grip_recording = False
        self.hints_recording = False
        self.session_recording = False

        self.drag_active = False
        self.drag_start_time = None
        self.drag_times = {}
        self.drag_record_times = {}
        self.performance_summary_text = ""

        self._last_update = 0.0
        self._last_live_graph_update = 0.0
        self.live_graph_paused = bool(self.manager.config.get('live_graph_paused', False))

        self.build_ui()

    def build_ui(self):
        root = QVBoxLayout(self)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        root.addWidget(scroll)

        page = QWidget()
        scroll.setWidget(page)
        outer = QVBoxLayout(page)

        title = QLabel("PERFORMANCE LAB PROTOTYPE")
        title.setStyleSheet("font-size: 18px; font-weight: 800; color: #00d9ff; letter-spacing: 1px;")
        outer.addWidget(title)

        subtitle = QLabel("Internal test area: Live Graph, Drag Timer, Grip Monitor, Smart Hints, Session Report, HUD Presets, Profiles and Support Info.")
        subtitle.setWordWrap(True)
        outer.addWidget(subtitle)

        self.lbl_vehicle_live = QLabel("Vehicle: waiting for telemetry")
        self.lbl_vehicle_live.setWordWrap(True)
        self.lbl_vehicle_live.setStyleSheet("font-weight:800; color:#00d9ff; background:rgba(2,8,14,130); border:1px solid rgba(0,217,255,80); border-radius:9px; padding:8px;")
        outer.addWidget(self.lbl_vehicle_live)

        self.build_live_graph_section(outer)
        self.build_drag_section(outer)
        self.build_grip_section(outer)
        self.build_hints_section(outer)
        self.build_session_section(outer)
        self.build_performance_summary_section(outer)
        self.build_presets_section(outer)
        self.build_profiles_section(outer)
        self.build_support_section(outer)

        outer.addStretch(1)

    def _section_style(self):
        return "QGroupBox{font-weight:700; color:#d9f7ff; border:1px solid rgba(0,217,255,90); border-radius:10px; margin-top:10px; padding:10px;}"

    def build_live_graph_section(self, outer):
        box = QGroupBox("Live Telemetry Graph")
        box.setStyleSheet(self._section_style())
        layout = QVBoxLayout(box)
        outer.addWidget(box)

        hint = QLabel("Time-based live graph. Separate from DynoClean RPM graph.")
        hint.setWordWrap(True)
        layout.addWidget(hint)

        row = QHBoxLayout()
        layout.addLayout(row)
        self.btn_live_graph_pause = QPushButton("Resume Live Graph" if self.live_graph_paused else "Pause Live Graph")
        self.btn_live_graph_pause.clicked.connect(self.pause_resume_live_graph)
        row.addWidget(self.btn_live_graph_pause)

        self.btn_live_graph_clear = QPushButton("Clear Live Graph")
        self.btn_live_graph_clear.clicked.connect(self.clear_live_graph)
        row.addWidget(self.btn_live_graph_clear)

        self.btn_live_graph_toggle = QPushButton("Show/Hide Graph")
        self.btn_live_graph_toggle.clicked.connect(self.pause_resume_live_graph_visible)
        row.addWidget(self.btn_live_graph_toggle)
        row.addStretch(1)

        self.live_graph = LiveGraphWidget()
        self.live_graph.set_config(self.manager.config)
        layout.addWidget(self.live_graph, 1)

    def build_drag_section(self, outer):
        box = QGroupBox("Drag Timer")
        box.setStyleSheet(self._section_style())
        layout = QVBoxLayout(box)
        outer.addWidget(box)

        self.lbl_drag_live = QLabel("Live Drag Timer: active · waiting for telemetry")
        self.lbl_drag_record = QLabel("Drag Record: stopped")
        for lab in [self.lbl_drag_live, self.lbl_drag_record]:
            lab.setWordWrap(True)
            lab.setStyleSheet("font-weight:bold; color:#d9f7ff;")
            layout.addWidget(lab)

        row = QHBoxLayout()
        layout.addLayout(row)

        self.btn_drag_toggle = QPushButton("Start Drag Record")
        self.btn_drag_toggle.clicked.connect(self.toggle_drag_record)
        row.addWidget(self.btn_drag_toggle)

        self.btn_drag_reset = QPushButton("Reset Drag")
        self.btn_drag_reset.clicked.connect(self.reset_drag)
        row.addWidget(self.btn_drag_reset)

        row.addStretch(1)

    def build_grip_section(self, outer):
        box = QGroupBox("Grip Monitor")
        box.setStyleSheet(self._section_style())
        layout = QVBoxLayout(box)
        outer.addWidget(box)

        self.lbl_grip_live = QLabel("Grip Live Monitor: active · waiting for telemetry")
        self.lbl_grip_record = QLabel("Grip Record: stopped")
        self.lbl_grip_analysis = QLabel("Grip Analysis: no record yet")
        for lab in [self.lbl_grip_live, self.lbl_grip_record, self.lbl_grip_analysis]:
            lab.setWordWrap(True)
            lab.setStyleSheet("font-weight:bold; color:#d9f7ff;")
            layout.addWidget(lab)

        row = QHBoxLayout()
        layout.addLayout(row)

        self.btn_grip_toggle = QPushButton("Start Grip Record")
        self.btn_grip_toggle.clicked.connect(self.toggle_grip_record)
        row.addWidget(self.btn_grip_toggle)

        self.btn_grip_analyze = QPushButton("Analyze Grip")
        self.btn_grip_analyze.clicked.connect(self.analyze_grip)
        row.addWidget(self.btn_grip_analyze)

        self.btn_grip_reset = QPushButton("Reset Grip")
        self.btn_grip_reset.clicked.connect(self.reset_grip)
        row.addWidget(self.btn_grip_reset)

        row.addStretch(1)

    def build_hints_section(self, outer):
        box = QGroupBox("Smart Hints")
        box.setStyleSheet(self._section_style())
        layout = QVBoxLayout(box)
        outer.addWidget(box)

        self.lbl_hints_live = QLabel("Smart Hints Live: active · waiting for telemetry")
        self.lbl_hints_record = QLabel("Smart Hints Record: stopped")
        self.lbl_hints_analysis = QLabel("Smart Hints Analysis: no record yet")
        for lab in [self.lbl_hints_live, self.lbl_hints_record, self.lbl_hints_analysis]:
            lab.setWordWrap(True)
            lab.setStyleSheet("font-weight:bold; color:#d9f7ff;")
            layout.addWidget(lab)

        row = QHBoxLayout()
        layout.addLayout(row)

        self.btn_hints_toggle = QPushButton("Start Smart Hints Record")
        self.btn_hints_toggle.clicked.connect(self.toggle_hints_record)
        row.addWidget(self.btn_hints_toggle)

        self.btn_hints_analyze = QPushButton("Analyze Smart Hints")
        self.btn_hints_analyze.clicked.connect(self.analyze_hints)
        row.addWidget(self.btn_hints_analyze)

        self.btn_hints_reset = QPushButton("Reset Smart Hints")
        self.btn_hints_reset.clicked.connect(self.reset_hints)
        row.addWidget(self.btn_hints_reset)

        row.addStretch(1)

    def build_session_section(self, outer):
        box = QGroupBox("Session Report")
        box.setStyleSheet(self._section_style())
        layout = QVBoxLayout(box)
        outer.addWidget(box)

        self.lbl_session_live = QLabel("Session Live Stats: active · waiting for telemetry")
        self.lbl_session_record = QLabel("Session Record: stopped")
        for lab in [self.lbl_session_live, self.lbl_session_record]:
            lab.setWordWrap(True)
            lab.setStyleSheet("font-weight:bold; color:#d9f7ff;")
            layout.addWidget(lab)

        row = QHBoxLayout()
        layout.addLayout(row)

        self.btn_session_toggle = QPushButton("Start Session Record")
        self.btn_session_toggle.clicked.connect(self.toggle_session_record)
        row.addWidget(self.btn_session_toggle)

        self.btn_session_export = QPushButton("Export Session Report")
        self.btn_session_export.clicked.connect(self.export_session_report)
        row.addWidget(self.btn_session_export)

        self.btn_session_reset = QPushButton("Reset Session")
        self.btn_session_reset.clicked.connect(self.reset_session)
        row.addWidget(self.btn_session_reset)

        row.addStretch(1)

    def build_performance_summary_section(self, outer):
        box = QGroupBox("Vehicle Analysis / Performance Summary")
        box.setStyleSheet(self._section_style())
        layout = QVBoxLayout(box)
        outer.addWidget(box)

        hint = QLabel("Local telemetry-based summary. ONYX only reads telemetry values; it does not know your installed tuning parts.")
        hint.setWordWrap(True)
        layout.addWidget(hint)

        self.lbl_performance_summary = QLabel("No analysis yet. Run telemetry, then press Analyze Current Session.")
        self.lbl_performance_summary.setWordWrap(True)
        self.lbl_performance_summary.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        self.lbl_performance_summary.setStyleSheet(
            "font-weight:600; color:#d9f7ff; background:rgba(2,8,14,120); "
            "border:1px solid rgba(0,217,255,80); border-radius:10px; padding:10px;"
        )
        layout.addWidget(self.lbl_performance_summary)

        row = QHBoxLayout()
        layout.addLayout(row)

        self.btn_analyze_summary = QPushButton("Analyze Current Session")
        self.btn_analyze_summary.clicked.connect(self.analyze_performance_summary)
        row.addWidget(self.btn_analyze_summary)

        self.btn_copy_summary = QPushButton("Copy Summary")
        self.btn_copy_summary.clicked.connect(self.copy_performance_summary)
        row.addWidget(self.btn_copy_summary)

        self.btn_reset_summary = QPushButton("Reset Summary")
        self.btn_reset_summary.clicked.connect(self.reset_performance_summary)
        row.addWidget(self.btn_reset_summary)

        row.addStretch(1)

    def build_presets_section(self, outer):
        box = QGroupBox("HUD Presets")
        box.setStyleSheet(self._section_style())
        row = QHBoxLayout(box)
        outer.addWidget(box)

        for name in ["Minimal", "Race", "Dyno", "Drag", "Tuning", "Streamer"]:
            b = QPushButton(name)
            b.clicked.connect(lambda _, n=name: self.apply_hud_preset(n))
            row.addWidget(b)
        row.addStretch(1)

    def build_profiles_section(self, outer):
        box = QGroupBox("Profiles")
        box.setStyleSheet(self._section_style())
        row = QHBoxLayout(box)
        outer.addWidget(box)

        self.profile_name = QLineEdit(self.manager.config.get("active_profile", "Default"))
        row.addWidget(QLabel("Profile Name:"))
        row.addWidget(self.profile_name)

        self.btn_save_profile = QPushButton("Save Profile")
        self.btn_save_profile.clicked.connect(self.save_profile)
        row.addWidget(self.btn_save_profile)

        self.btn_load_profile = QPushButton("Load Profile")
        self.btn_load_profile.clicked.connect(self.load_profile)
        row.addWidget(self.btn_load_profile)

    def build_support_section(self, outer):
        box = QGroupBox("Support Info")
        box.setStyleSheet(self._section_style())
        layout = QVBoxLayout(box)
        outer.addWidget(box)

        row = QHBoxLayout()
        layout.addLayout(row)

        self.btn_support = QPushButton("Copy Support Info")
        self.btn_support.clicked.connect(self.copy_support_info)
        row.addWidget(self.btn_support)

        self.btn_open_crash_folder = QPushButton("Open Crash Folder")
        self.btn_open_crash_folder.clicked.connect(self.open_crash_folder)
        row.addWidget(self.btn_open_crash_folder)

        self.btn_clear_log = QPushButton("Clear Crash Log")
        self.btn_clear_log.clicked.connect(self.clear_crash_log)
        row.addWidget(self.btn_clear_log)

        row.addStretch(1)

        self.lbl_support = QLabel(f"Crash log: {CRASH_LOG_PATH}")
        self.lbl_support.setWordWrap(True)
        layout.addWidget(self.lbl_support)

    # ---------- telemetry ----------
    def add_sample(self, t):
        try:
            if t is None:
                return
            if hasattr(self, "lbl_vehicle_live"):
                self.lbl_vehicle_live.setText(vehicle_summary_text(t))

            self.samples.append(t)
            if len(self.samples) > 12000:
                self.samples = self.samples[-12000:]

            if self.drag_recording:
                self.drag_samples.append(t)
                self.drag_samples = self.drag_samples[-12000:]

            if self.grip_recording:
                self.grip_samples.append(t)
                self.grip_samples = self.grip_samples[-12000:]

            if self.hints_recording:
                self.hints_samples.append(t)
                self.hints_samples = self.hints_samples[-12000:]

            if self.session_recording:
                self.session_samples.append(t)
                self.session_samples = self.session_samples[-12000:]

            self.update_drag_live(t)

            now = time.time()
            if now - self._last_update >= 0.10:
                self._last_update = now
                self.update_all_labels()

            if (not self.live_graph_paused) and now - self._last_live_graph_update >= 0.10:
                self._last_live_graph_update = now
                self.live_graph.set_config(self.manager.config)
                self.live_graph.set_samples(self.samples)
        except Exception as exc:
            log_error("PerformanceLabTab.add_sample", exc)

    # ---------- drag ----------
    def update_drag_live(self, t):
        try:
            speed_kmh = getattr(t, "speed_kmh", 0.0)
            throttle = getattr(t, "throttle_pct", 0.0)

            if speed_kmh < 2 and throttle < 10:
                self.drag_active = False
                self.drag_start_time = None
                # Do not clear drag_times constantly if it already has a result; start clears on next launch.

            if not self.drag_active and speed_kmh >= 5 and throttle > 50:
                self.drag_active = True
                self.drag_start_time = getattr(t, "timestamp", time.time())
                self.drag_times = {}

            if self.drag_active and self.drag_start_time:
                for low, high, key in [(0,100,"0-100"),(0,200,"0-200"),(100,200,"100-200"),(200,300,"200-300")]:
                    if key not in self.drag_times:
                        if key.startswith("0-"):
                            if speed_kmh >= high:
                                self.drag_times[key] = getattr(t, "timestamp", time.time()) - self.drag_start_time
                        else:
                            low_key = f"0-{low}"
                            if low_key in self.drag_times and speed_kmh >= high:
                                self.drag_times[key] = getattr(t, "timestamp", time.time()) - (self.drag_start_time + self.drag_times[low_key])

            self.manager.drag_overlay_times = dict(self.drag_times)
            self.manager.drag_overlay_recording = bool(self.drag_recording)
            if self.manager.overlay:
                self.manager.overlay.update()
        except Exception as exc:
            log_error("PerformanceLabTab.update_drag_live", exc)

    def toggle_drag_record(self):
        try:
            self.drag_recording = not self.drag_recording
            if self.drag_recording:
                self.drag_samples = []
                self.drag_record_times = {}
                self.btn_drag_toggle.setText("Stop Drag Record")
                self.lbl_drag_record.setText("Drag Record: recording...")
            else:
                self.btn_drag_toggle.setText("Start Drag Record")
                self.compute_drag_record_times()
                self.manager.drag_overlay_record_times = dict(self.drag_record_times)
                self.manager.drag_overlay_recording = False
                self.update_all_labels()
                if self.manager.overlay:
                    self.manager.overlay.update()
        except Exception as exc:
            log_error("PerformanceLabTab.toggle_drag_record", exc)

    def compute_drag_record_times(self):
        try:
            self.drag_record_times = self.calc_drag_times_for_samples(self.drag_samples)
        except Exception as exc:
            log_error("PerformanceLabTab.compute_drag_record_times", exc)

    def calc_drag_times_for_samples(self, samples):
        if not samples:
            return {}
        start_time = None
        times = {}
        zero100_time = None
        zero200_time = None

        for s in samples:
            speed = getattr(s, "speed_kmh", 0.0)
            throttle = getattr(s, "throttle_pct", 0.0)
            ts = getattr(s, "timestamp", time.time())
            if start_time is None and speed >= 5 and throttle > 40:
                start_time = ts
            if start_time is None:
                continue

            if "0-100" not in times and speed >= 100:
                times["0-100"] = ts - start_time
                zero100_time = ts
            if "0-200" not in times and speed >= 200:
                times["0-200"] = ts - start_time
                zero200_time = ts
            if "100-200" not in times and zero100_time and speed >= 200:
                times["100-200"] = ts - zero100_time
            if "200-300" not in times and zero200_time and speed >= 300:
                times["200-300"] = ts - zero200_time
        return times

    def reset_drag(self):
        self.drag_samples = []
        self.drag_times = {}
        self.drag_record_times = {}
        self.drag_recording = False
        self.drag_active = False
        self.drag_start_time = None
        self.manager.drag_overlay_times = {}
        self.manager.drag_overlay_record_times = {}
        self.manager.drag_overlay_recording = False
        self.btn_drag_toggle.setText("Start Drag Record")
        self.update_all_labels()
        if self.manager.overlay:
            self.manager.overlay.update()

    # ---------- grip ----------
    def grip_status_for(self, s):
        front = getattr(s, "front_combined_slip", 0.0)
        rear = getattr(s, "rear_combined_slip", 0.0)
        f_ratio = getattr(s, "front_slip_ratio", 0.0)
        r_ratio = getattr(s, "rear_slip_ratio", 0.0)
        speed = getattr(s, "speed_kmh", 0.0)
        throttle = getattr(s, "throttle_pct", 0.0)
        steer = abs(getattr(s, "steer", 0))

        status = "GRIP OK"
        severity = 0

        if rear > 1.6 or r_ratio > 0.65:
            status, severity = "REAR SLIP / OVERSTEER RISK", 3
        elif front > 1.6 or f_ratio > 0.65:
            status, severity = "FRONT SLIP / UNDERSTEER RISK", 3
        elif rear > 0.9 and rear > front * 1.25:
            status, severity = "REAR SLIP WARNING", 2
        elif front > 0.9 and front > rear * 1.25:
            status, severity = "FRONT SLIP WARNING", 2
        elif speed > 180 and steer > 70 and throttle > 80:
            status, severity = "HIGH LOAD WARNING", 2

        return status, severity, front, rear, f_ratio, r_ratio

    def toggle_grip_record(self):
        self.grip_recording = not self.grip_recording
        if self.grip_recording:
            self.grip_samples = []
            self.btn_grip_toggle.setText("Stop Grip Record")
            self.lbl_grip_record.setText("Grip Record: recording...")
        else:
            self.btn_grip_toggle.setText("Start Grip Record")
            self.analyze_grip()

    def analyze_grip(self):
        try:
            samples = self.grip_samples if self.grip_samples else self.samples[-1000:]
            text = self.build_grip_analysis(samples)
            self.lbl_grip_analysis.setText("Grip Analysis: " + text)
        except Exception as exc:
            log_error("PerformanceLabTab.analyze_grip", exc)

    def build_grip_analysis(self, samples):
        if not samples:
            return "no data"
        avg_front = sum(getattr(s, "front_combined_slip", 0.0) for s in samples) / max(1, len(samples))
        avg_rear = sum(getattr(s, "rear_combined_slip", 0.0) for s in samples) / max(1, len(samples))
        max_speed = max(getattr(s, "speed_kmh", 0.0) for s in samples)
        critical = [self.grip_status_for(s)[0] for s in samples if self.grip_status_for(s)[1] >= 3]

        if avg_front == 0 and avg_rear == 0:
            return "Slip telemetry may be unavailable or too low. Fallback uses speed/throttle/steering only."

        direction = "balanced"
        if avg_rear > avg_front * 1.2:
            direction = "rear is more unstable"
        elif avg_front > avg_rear * 1.2:
            direction = "front is more unstable"

        crit_txt = f" Critical events: {len(critical)}." if critical else " No critical events detected."
        return f"{direction}. Avg Front {avg_front:.2f}, Avg Rear {avg_rear:.2f}, Max Speed {max_speed:.1f} km/h.{crit_txt}"

    def reset_grip(self):
        self.grip_samples = []
        self.grip_recording = False
        self.btn_grip_toggle.setText("Start Grip Record")
        self.lbl_grip_record.setText("Grip Record: stopped")
        self.lbl_grip_analysis.setText("Grip Analysis: reset")

    # ---------- hints ----------
    def make_hints_for_samples(self, samples):
        if not samples:
            return []
        recent = samples[-500:]
        avg_rear = sum(getattr(s, "rear_combined_slip", 0.0) for s in recent) / max(1, len(recent))
        avg_front = sum(getattr(s, "front_combined_slip", 0.0) for s in recent) / max(1, len(recent))
        high_speed = max(getattr(s, "speed_kmh", 0.0) for s in recent)
        avg_throttle = sum(getattr(s, "throttle_pct", 0.0) for s in recent) / max(1, len(recent))
        hints = []

        if avg_rear > avg_front * 1.25 and avg_rear > 0.9:
            hints.append("Rear slip is higher than front: try softer rear ARB, lower rear tire pressure, or more rear stability.")
        if avg_front > avg_rear * 1.25 and avg_front > 0.9:
            hints.append("Front slip is higher than rear: possible understeer; adjust front stiffness/aero/tire pressure.")
        if high_speed > 250 and avg_rear > 0.8:
            hints.append("High-speed rear slip detected: add rear stability or reduce aggressive rear setup.")
        if avg_throttle > 75 and avg_rear > 1.2:
            hints.append("Throttle-on rear instability detected: check diff acceleration, rear tires or rear stiffness.")

        if avg_front == 0 and avg_rear == 0:
            last = recent[-1]
            if getattr(last, "speed_kmh", 0) > 180 and abs(getattr(last, "steer", 0)) > 70 and getattr(last, "throttle_pct", 0) > 80:
                hints.append("High speed + high steering + throttle: watch for aero/ARB/diff instability.")
        return hints[:4]

    def toggle_hints_record(self):
        self.hints_recording = not self.hints_recording
        if self.hints_recording:
            self.hints_samples = []
            self.btn_hints_toggle.setText("Stop Smart Hints Record")
            self.lbl_hints_record.setText("Smart Hints Record: recording...")
        else:
            self.btn_hints_toggle.setText("Start Smart Hints Record")
            self.analyze_hints()

    def analyze_hints(self):
        try:
            samples = self.hints_samples if self.hints_samples else self.samples[-1000:]
            hints = self.make_hints_for_samples(samples)
            self.lbl_hints_analysis.setText("Smart Hints Analysis: " + (" | ".join(hints) if hints else "No major issue detected."))
        except Exception as exc:
            log_error("PerformanceLabTab.analyze_hints", exc)

    def reset_hints(self):
        self.hints_samples = []
        self.hints_recording = False
        self.btn_hints_toggle.setText("Start Smart Hints Record")
        self.lbl_hints_record.setText("Smart Hints Record: stopped")
        self.lbl_hints_analysis.setText("Smart Hints Analysis: reset")

    # ---------- session ----------
    def toggle_session_record(self):
        self.session_recording = not self.session_recording
        if self.session_recording:
            self.session_samples = []
            self.btn_session_toggle.setText("Stop Session Record")
            self.lbl_session_record.setText("Session Record: recording...")
        else:
            self.btn_session_toggle.setText("Start Session Record")
            self.update_all_labels()

    def reset_session(self):
        self.session_samples = []
        self.session_recording = False
        self.btn_session_toggle.setText("Start Session Record")
        self.lbl_session_record.setText("Session Record: stopped")
        self.update_all_labels()

    # ---------- performance summary ----------
    def _pick_line(self, lines):
        try:
            return random.choice(lines)
        except Exception:
            return lines[0] if lines else ""

    def _fmt_time(self, val):
        return f"{val:.2f}s" if isinstance(val, (int, float)) else "-"

    def build_performance_summary_text(self, samples):
        if not samples:
            return "No telemetry samples available yet. Drive for a bit or start a session record first."

        cfg = self.manager.config
        source = samples[-12000:]

        peak_speed_s = max(source, key=lambda s: getattr(s, "speed_kmh", 0.0))
        peak_power_s = max(source, key=lambda s: power_value(s, cfg))
        peak_boost_s = max(source, key=lambda s: boost_value(s, cfg))
        peak_rpm_s = max(source, key=lambda s: getattr(s, "rpm", 0.0))

        peak_speed = speed_value(peak_speed_s, cfg)
        peak_power = power_value(peak_power_s, cfg)
        peak_boost = boost_value(peak_boost_s, cfg)
        peak_rpm = float(getattr(peak_rpm_s, "rpm", 0.0) or 0.0)

        speeds_kmh = [float(getattr(s, "speed_kmh", 0.0) or 0.0) for s in source]
        avg_speed_kmh = sum(speeds_kmh) / max(1, len(speeds_kmh))
        throttles = [float(getattr(s, "throttle_pct", 0.0) or 0.0) for s in source]
        avg_throttle = sum(throttles) / max(1, len(throttles))

        front_slips = [float(getattr(s, "front_combined_slip", 0.0) or 0.0) for s in source]
        rear_slips = [float(getattr(s, "rear_combined_slip", 0.0) or 0.0) for s in source]
        avg_front = sum(front_slips) / max(1, len(front_slips))
        avg_rear = sum(rear_slips) / max(1, len(rear_slips))
        max_front = max(front_slips) if front_slips else 0.0
        max_rear = max(rear_slips) if rear_slips else 0.0

        drag_times = dict(self.drag_times or {})
        if not drag_times:
            drag_times = self.calc_drag_times_for_samples(source)

        hints = self.make_hints_for_samples(source)
        grip_text = self.build_grip_analysis(source)

        car_id = None
        for attr in ["car_ordinal", "CarOrdinal", "carOrdinal"]:
            if hasattr(source[-1], attr):
                try:
                    car_id = getattr(source[-1], attr)
                    if car_id == 0:
                        car_id = None
                    break
                except Exception:
                    pass

        # Conservative local scoring. No claims about installed parts.
        score = 50
        if peak_speed > 250: score += 10
        if peak_speed > 320: score += 8
        if peak_power > 450: score += 10
        if peak_power > 800: score += 8
        if "0-100" in drag_times:
            if drag_times["0-100"] <= 3.5: score += 12
            elif drag_times["0-100"] <= 5.0: score += 7
            elif drag_times["0-100"] >= 8.0: score -= 12
        if "100-200" in drag_times:
            if drag_times["100-200"] <= 5.5: score += 10
            elif drag_times["100-200"] >= 10.0: score -= 10
        if avg_rear > 1.2 or avg_front > 1.2: score -= 10
        if max_rear > 2.3 or max_front > 2.3: score -= 8
        if avg_throttle > 70 and peak_speed < 180: score -= 8
        score = max(0, min(100, int(score)))

        if score >= 82:
            verdict_pool = [
                "Overall: Strong setup. The car looks quick and reasonably controlled.",
                "Overall: This run looks solid. The telemetry shows good pace without obvious chaos.",
                "Overall: The car is doing well. Most improvements from here are probably fine-tuning, not a full rebuild.",
                "Overall: This one actually cooks. The data shows a strong baseline."
            ]
        elif score >= 65:
            verdict_pool = [
                "Overall: Good baseline. The car performs well, but there is still room for cleaner execution.",
                "Overall: The setup looks usable and competitive, with a few areas worth checking.",
                "Overall: Not bad at all. The car has pace, but some telemetry hints suggest polish work.",
                "Overall: The car looks decent. Nothing screams broken, but it can still be sharpened."
            ]
        elif score >= 45:
            verdict_pool = [
                "Overall: Mixed run. The car has usable performance, but the data shows wasted potential.",
                "Overall: The car is not terrible, but the telemetry does not call it properly dialed in yet.",
                "Overall: This looks like a work-in-progress setup. It moves, but it does not look fully clean.",
                "Overall: The setup has potential, but the run shows some efficiency problems."
            ]
        else:
            verdict_pool = [
                "Overall: Weak run. The car is trying, but the telemetry says it needs setup work.",
                "Overall: This setup is fighting itself. There is too much input for too little clean result.",
                "Overall: The car is usable, but the data does not look competitive yet.",
                "Overall: This looks more like a learning run than a winning run."
            ]

        positives = []
        if peak_power > 500:
            positives.append(self._pick_line([
                "Power output looks healthy based on the recorded peak.",
                "The car has enough output to make the run interesting.",
                "Peak power is not the obvious weak point here."
            ]))
        if peak_speed > 250:
            positives.append(self._pick_line([
                "Top speed behavior looks respectable.",
                "The car is capable of building serious speed.",
                "High-speed potential is present."
            ]))
        if avg_rear < 0.8 and avg_front < 0.8:
            positives.append(self._pick_line([
                "Grip behavior looks controlled in the recorded data.",
                "The car does not appear to waste much power through slip.",
                "Traction looks reasonably stable during this sample."
            ]))
        if not hints:
            positives.append("No major telemetry warning stood out during the analyzed sample.")

        issues = []
        if avg_rear > avg_front * 1.25 and avg_rear > 0.9:
            issues.append("Rear slip is higher than front slip. The car may be traction-limited under throttle.")
        if avg_front > avg_rear * 1.25 and avg_front > 0.9:
            issues.append("Front slip is higher than rear slip. The car may be pushing wide or overloading the front.")
        if max_rear > 2.3:
            issues.append("Critical rear slip spikes were detected. That can hurt launch and acceleration consistency.")
        if max_front > 2.3:
            issues.append("Critical front slip spikes were detected. That can make the car less predictable.")
        if avg_throttle > 70 and peak_speed < 180:
            issues.append("Throttle input was high, but speed gain looked limited. This may point to grip, gearing, or test conditions.")
        if "0-100" in drag_times and drag_times["0-100"] > 6.5:
            issues.append("0-100 time looks slow for an acceleration test. Launch grip, gearing, or throttle control may be limiting it.")
        if "100-200" in drag_times and drag_times["100-200"] > 9.5:
            issues.append("100-200 pull looks weak. Mid-range acceleration may need attention.")

        if not issues:
            issues.append("No major problem was detected. Any changes should be small and tested one at a time.")

        suggestions = []
        if avg_rear > 1.0:
            suggestions.append("If traction is the goal, check rear grip balance, differential behavior, tire pressure, and launch input.")
        if avg_front > 1.0:
            suggestions.append("If the car pushes wide, check front grip balance, tire pressure, suspension stiffness, or aero balance.")
        if "100-200" in drag_times and drag_times["100-200"] > 9.5:
            suggestions.append("For better mid-range pull, check gearing before assuming the car needs more parts.")
        if not suggestions:
            suggestions.append("If you want to improve it further, focus on fine-tuning: gearing, tire pressure, differential, suspension, and repeatable test runs.")
        suggestions.append("ONYX only sees telemetry. It does not know which engine parts or upgrades are installed, so treat this as tuning feedback, not a parts shopping list.")

        funny_line = self._pick_line([
            "Telemetry mood: the car is either cooking or politely asking for another tuning pass.",
            "ONYX verdict style: not magic, just telemetry being brutally honest.",
            "The graph does not lie, but it also does not know your full build sheet.",
            "If this was not a clean straight run, retest before bullying the car too hard.",
            "This is tuning territory: change one thing, test again, then blame the car."
        ])

        lines = []
        lines.append("VEHICLE ANALYSIS / PERFORMANCE SUMMARY")
        lines.append("=" * 48)
        lines.append(self._pick_line(verdict_pool))
        lines.append("")
        lines.append(f"Telemetry Score: {score}/100")
        if car_id is not None:
            lines.append(f"Detected Car ID: {car_id}")
            vehicle_name = vehicle_name_for_id(car_id)
            if vehicle_name:
                lines.append(f"Detected Vehicle: {vehicle_name}")
            else:
                lines.append("Detected Vehicle: unknown (add mapping to car_database.json)")
        else:
            raw_ordinal = getattr(source[-1], "official_car_ordinal_raw", None)
            lines.append("Detected Car ID: unavailable")
            lines.append(f"Official CarOrdinal raw: {raw_ordinal}")
            lines.append("Detected Vehicle: unavailable")
            probe = summarize_vehicle_id_probe(source)
            if probe:
                lines.append("Vehicle ID probe candidates, compare after switching car:")
                for item in probe[:5]:
                    lines.append(f"- offset {item.get('offset')}: {item.get('value')} seen {item.get('seen')}x")
            else:
                lines.append("Vehicle ID probe: no stable alternative candidates found")
        lines.append(f"Samples analyzed: {len(source)}")
        lines.append("")
        lines.append("Peaks:")
        lines.append(f"- Peak Speed: {peak_speed:.1f} {speed_label(cfg)}")
        lines.append(f"- Peak RPM: {peak_rpm:.0f}")
        lines.append(f"- Peak {power_label(cfg)}: {peak_power:.1f}")
        lines.append(f"- Peak Boost: {peak_boost:.2f} {boost_label(cfg)}")
        lines.append("")
        lines.append("Drag:")
        for key in ["0-100", "0-200", "100-200", "200-300"]:
            lines.append(f"- {key}: {self._fmt_time(drag_times.get(key))}")
        lines.append("")
        lines.append("What looks good:")
        for p in positives[:4] or ["The car produced usable telemetry data for analysis."]:
            lines.append(f"- {p}")
        lines.append("")
        lines.append("What needs attention:")
        for issue in issues[:5]:
            lines.append(f"- {issue}")
        lines.append("")
        lines.append("Grip summary:")
        lines.append(f"- {grip_text}")
        lines.append("")
        lines.append("Suggestions:")
        for s in suggestions[:5]:
            lines.append(f"- {s}")
        if hints:
            lines.append("")
            lines.append("Smart hints:")
            for h in hints[:4]:
                lines.append(f"- {h}")
        lines.append("")
        lines.append(f"Note: {funny_line}")

        return "\n".join(lines)

    def analyze_performance_summary(self):
        try:
            samples = self.session_samples if self.session_samples else self.samples
            text = self.build_performance_summary_text(samples)
            self.performance_summary_text = text
            self.lbl_performance_summary.setText(text)
        except Exception as exc:
            log_error("PerformanceLabTab.analyze_performance_summary", exc)
            QMessageBox.warning(self, "Error", "Could not analyze current session.")

    def copy_performance_summary(self):
        try:
            text = self.performance_summary_text or self.lbl_performance_summary.text()
            QApplication.clipboard().setText(text)
            QMessageBox.information(self, "Copied", "Performance summary copied to clipboard.")
        except Exception as exc:
            log_error("PerformanceLabTab.copy_performance_summary", exc)

    def reset_performance_summary(self):
        self.performance_summary_text = ""
        if hasattr(self, "lbl_performance_summary"):
            self.lbl_performance_summary.setText("No analysis yet. Run telemetry, then press Analyze Current Session.")

    def export_session_report(self):
        try:
            samples = self.session_samples if self.session_samples else self.samples
            if not samples:
                QMessageBox.information(self, "No data", "No telemetry samples recorded yet.")
                return
            out_dir = Path("reports")
            out_dir.mkdir(exist_ok=True)
            stamp = time.strftime("%Y%m%d_%H%M%S")
            path = out_dir / f"onyx_performance_lab_report_{stamp}.txt"
            cfg = self.manager.config
            peak_speed = max(samples, key=lambda s: getattr(s, "speed_kmh", 0.0))
            peak_power = max(samples, key=lambda s: power_value(s, cfg))
            peak_boost = max(samples, key=lambda s: boost_value(s, cfg))
            peak_rpm = max(samples, key=lambda s: getattr(s, "rpm", 0.0))
            hints = self.make_hints_for_samples(samples)
            grip_text = self.build_grip_analysis(samples)
            drag_times = self.calc_drag_times_for_samples(samples)

            with path.open("w", encoding="utf-8") as f:
                f.write("ONYX Drive HUD v5.2.7e Vehicle Database No-AI Report\n")
                f.write("="*70 + "\n")
                f.write(f"Samples: {len(samples)}\n")
                f.write(f"Peak Speed: {speed_value(peak_speed,cfg):.2f} {speed_label(cfg)}\n")
                f.write(f"Peak RPM: {getattr(peak_rpm,'rpm',0):.0f}\n")
                f.write(f"Peak {power_label(cfg)}: {power_value(peak_power,cfg):.2f}\n")
                f.write(f"Peak Boost: {boost_value(peak_boost,cfg):.2f} {boost_label(cfg)}\n")
                f.write("\nDrag Times:\n")
                for key in ["0-100","0-200","100-200","200-300"]:
                    val = drag_times.get(key)
                    f.write(f"- {key}: {val:.2f}s\n" if isinstance(val, (int,float)) else f"- {key}: -\n")
                f.write("\nGrip Analysis:\n")
                f.write(grip_text + "\n")
                f.write("\nSmart Hints:\n")
                for h in hints or ["No major issue detected."]:
                    f.write(f"- {h}\n")
            QMessageBox.information(self, "Report exported", str(path))
        except Exception as exc:
            log_error("PerformanceLabTab.export_session_report", exc)
            QMessageBox.warning(self, "Error", "Could not export session report.")

    # ---------- labels ----------
    def update_all_labels(self):
        try:
            if not self.samples:
                return
            cfg = self.manager.config
            last = self.samples[-1]

            drag_parts = []
            for key in ["0-100", "0-200", "100-200", "200-300"]:
                v = self.drag_times.get(key)
                drag_parts.append(f"{key}: {v:.2f}s" if isinstance(v, (int,float)) else f"{key}: -")
            self.lbl_drag_live.setText("Live Drag Timer: active · " + " · ".join(drag_parts))
            self.manager.drag_overlay_times = dict(self.drag_times)
            self.manager.drag_overlay_recording = bool(self.drag_recording)

            record_drag = self.drag_record_times if self.drag_record_times else self.calc_drag_times_for_samples(self.drag_samples)
            drag_rec_parts = []
            for key in ["0-100", "0-200", "100-200", "200-300"]:
                v = record_drag.get(key)
                drag_rec_parts.append(f"{key}: {v:.2f}s" if isinstance(v, (int,float)) else f"{key}: -")
            self.lbl_drag_record.setText(("Drag Record: recording · " if self.drag_recording else "Drag Record: stopped · ") + " · ".join(drag_rec_parts))

            status, sev, front, rear, fr, rr = self.grip_status_for(last)
            self.lbl_grip_live.setText(f"Grip Live Monitor: active · {status} · Front {front:.2f} / Rear {rear:.2f} · Ratio F {fr:.2f} / R {rr:.2f}")
            self.lbl_grip_record.setText(f"Grip Record: {'recording' if self.grip_recording else 'stopped'} · Samples {len(self.grip_samples)}")

            hints = self.make_hints_for_samples(self.samples)
            self.lbl_hints_live.setText("Smart Hints Live: active · " + (" | ".join(hints) if hints else "No major issue detected."))
            self.lbl_hints_record.setText(f"Smart Hints Record: {'recording' if self.hints_recording else 'stopped'} · Samples {len(self.hints_samples)}")

            session_source = self.session_samples if self.session_samples else self.samples
            peak_speed = max(session_source, key=lambda s: getattr(s, "speed_kmh", 0.0))
            peak_power = max(session_source, key=lambda s: power_value(s, cfg))
            peak_boost = max(session_source, key=lambda s: boost_value(s, cfg))
            peak_rpm = max(session_source, key=lambda s: getattr(s, "rpm", 0.0))
            avg_throttle = sum(getattr(s, "throttle_pct", 0.0) for s in session_source[-500:]) / max(1, len(session_source[-500:]))

            self.lbl_session_live.setText(
                f"Session Live Stats: active · Samples {len(self.samples)} · "
                f"Peak Speed {speed_value(peak_speed,cfg):.1f} {speed_label(cfg)} · "
                f"Peak RPM {getattr(peak_rpm,'rpm',0):.0f} · "
                f"Peak {power_label(cfg)} {power_value(peak_power,cfg):.0f} · "
                f"Peak Boost {boost_value(peak_boost,cfg):.2f} {boost_label(cfg)} · "
                f"Avg Throttle {avg_throttle:.0f}%"
            )
            self.lbl_session_record.setText(f"Session Record: {'recording' if self.session_recording else 'stopped'} · Samples {len(self.session_samples)}")
        except Exception as exc:
            log_error("PerformanceLabTab.update_all_labels", exc)

    # ---------- live graph ----------
    def clear_live_graph(self):
        try:
            self.samples = []
            self.live_graph.set_samples([])
            self.update_all_labels()
        except Exception as exc:
            log_error("PerformanceLabTab.clear_live_graph", exc)

    def pause_resume_live_graph(self):
        try:
            self.live_graph_paused = not self.live_graph_paused
            self.manager.config["live_graph_paused"] = self.live_graph_paused
            save_config(self.manager.config)
            if hasattr(self, "btn_live_graph_pause"):
                self.btn_live_graph_pause.setText("Resume Live Graph" if self.live_graph_paused else "Pause Live Graph")
        except Exception as exc:
            log_error("PerformanceLabTab.pause_resume_live_graph", exc)

    def pause_resume_live_graph_visible(self):
        try:
            self.live_graph.setVisible(not self.live_graph.isVisible())
        except Exception as exc:
            log_error("PerformanceLabTab.pause_resume_live_graph_visible", exc)

    def toggle_live_graph_visible(self):
        return self.pause_resume_live_graph_visible()

    # Compatibility wrapper for older hotkey/config names.
    def toggle_live_graph(self):
        self.pause_resume_live_graph()

    # ---------- support/profile/presets ----------
    def copy_support_info(self):
        try:
            cfg = self.manager.config
            txt = (
                "ONYX Support Info\n"
                f"Version: ONYX Drive HUD v5.2.6\n"
                f"UDP: {cfg.get('udp_host','0.0.0.0')}:{cfg.get('udp_port',5607)}\n"
                f"Units: {cfg.get('unit_system')} · {cfg.get('speed_unit')} · {cfg.get('power_unit')} · {cfg.get('boost_unit')}\n"
                f"Profile: {cfg.get('active_profile','Default')}\n"
                f"Overlay Monitor Index: {cfg.get('overlay_monitor_index', -1)}\n"
                f"Last telemetry: {'yes' if self.manager.latest else 'no'}\n"
                f"Performance samples: {len(self.samples)}\n"
                f"Drag samples: {len(self.drag_samples)}\n"
                f"Grip samples: {len(self.grip_samples)}\n"
                f"Hints samples: {len(self.hints_samples)}\n"
                f"Session samples: {len(self.session_samples)}\n"
                f"Crash log: {CRASH_LOG_PATH}\n"
            )
            QApplication.clipboard().setText(txt)
            QMessageBox.information(self, "Copied", "Support info copied to clipboard.")
        except Exception as exc:
            log_error("PerformanceLabTab.copy_support_info", exc)

    def open_crash_folder(self):
        try:
            LOG_DIR.mkdir(exist_ok=True)
            path = LOG_DIR.resolve()
            if sys.platform.startswith("win"):
                os.startfile(str(path))
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(path)])
            else:
                subprocess.Popen(["xdg-open", str(path)])
        except Exception as exc:
            log_error("PerformanceLabTab.open_crash_folder", exc)
            QMessageBox.warning(self, "Error", "Could not open crash folder.")

    def clear_crash_log(self):
        try:
            if CRASH_LOG_PATH.exists():
                CRASH_LOG_PATH.unlink()
            QMessageBox.information(self, "OK", "Crash log cleared.")
        except Exception as exc:
            log_error("PerformanceLabTab.clear_crash_log", exc)

    def _profile_path(self):
        name = self.profile_name.text().strip() or "Default"
        safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", name)
        d = Path("profiles")
        d.mkdir(exist_ok=True)
        return d / f"{safe}.json"

    def save_profile(self):
        try:
            self.manager.collect_forms()
            self.manager.config["active_profile"] = self.profile_name.text().strip() or "Default"
            path = self._profile_path()
            path.write_text(json.dumps(self.manager.config, indent=2, ensure_ascii=False), encoding="utf-8")
            QMessageBox.information(self, "Profile saved", str(path))
        except Exception as exc:
            log_error("PerformanceLabTab.save_profile", exc)

    def load_profile(self):
        try:
            path, _ = QFileDialog.getOpenFileName(self, "Load ONYX profile", "profiles", "JSON (*.json)")
            if not path:
                return
            data = json.loads(Path(path).read_text(encoding="utf-8"))
            self.manager.config.update(data)
            save_config(self.manager.config)
            self.manager.rebuild_ui()
        except Exception as exc:
            log_error("PerformanceLabTab.load_profile", exc)
            QMessageBox.warning(self, "Error", "Could not load profile.")

    def apply_hud_preset(self, name):
        try:
            cards = self.manager.config["cards"]
            presets = {
                "Minimal": {
                    "speed": (35,85,190,76,True), "rpm": (35,168,190,76,True), "gear": (35,251,150,76,True),
                    "power": (35,334,180,76,False), "boost": (35,417,180,76,False), "grip": (35,500,190,76,True), "tachometer": (300,90,240,240,False), "drag_timer": (300,350,260,150,False)
                },
                "Race": {
                    "speed": (35,85,230,88,True), "rpm": (35,180,230,88,True), "gear": (35,275,230,88,True),
                    "power": (35,370,230,88,True), "boost": (35,465,230,88,True), "grip": (35,560,230,88,True), "tachometer": (300,90,260,260,False), "drag_timer": (300,370,260,150,False)
                },
                "Dyno": {
                    "speed": (35,85,210,78,True), "rpm": (35,170,210,78,True), "gear": (35,255,150,78,False),
                    "power": (35,340,210,78,True), "boost": (35,425,210,78,True), "grip": (35,510,210,78,True), "tachometer": (300,90,260,260,False), "drag_timer": (300,370,260,150,False)
                },
                "Drag": {
                    "speed": (40,80,260,100,True), "rpm": (40,190,220,80,True), "gear": (40,280,160,80,True),
                    "power": (40,370,220,80,True), "boost": (40,460,220,80,True), "grip": (40,550,220,80,True), "tachometer": (330,90,270,270,False), "drag_timer": (330,380,270,150,False)
                },
                "Tuning": {
                    "speed": (35,85,210,76,True), "rpm": (35,166,210,76,True), "gear": (35,247,160,76,True),
                    "power": (35,328,210,76,True), "boost": (35,409,210,76,True), "grip": (35,490,210,76,True), "tachometer": (300,90,260,260,False), "drag_timer": (300,370,260,150,False)
                },
                "Streamer": {
                    "speed": (50,120,260,105,True), "rpm": (50,235,260,105,True), "gear": (50,350,190,105,True),
                    "power": (50,465,260,105,True), "boost": (50,580,260,105,True), "grip": (50,695,260,105,True), "tachometer": (340,120,300,300,False), "drag_timer": (340,440,300,170,False)
                },
            }
            for key, vals in presets.get(name, presets["Race"]).items():
                x,y,w,h,vis = vals
                # Presets should only move/resize/show tiles.
                # Keep custom labels/colors so Save Layout or presets do not wipe user edits.
                if key not in cards:
                    cards[key] = json.loads(json.dumps(DEFAULT_CONFIG["cards"][key]))
                cards[key].update({"x":x,"y":y,"w":w,"h":h,"visible":vis})
            save_config(self.manager.config)
            if self.manager.overlay:
                self.manager.overlay.sync_config()
            QMessageBox.information(self, "Preset applied", f"{name} HUD preset applied.")
        except Exception as exc:
            log_error("PerformanceLabTab.apply_hud_preset", exc)



class ManagerWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config = load_config()
        self.queue = Queue(maxsize=3000)
        self.receiver = UdpReceiver(self.config["udp_host"], self.config["udp_port"], self.queue)
        self.receiver.start()
        self.latest = None
        self.drag_overlay_times = {}
        self.drag_overlay_record_times = {}
        self.drag_overlay_recording = False
        self.overlay = None
        self.vehicle_status_labels = []
        self.logged_unknown_vehicle_ids = set()
        self.setWindowTitle(tr(self.lang(), "title"))
        if ICON_PATH.exists():
            self.setWindowIcon(QIcon(str(ICON_PATH)))
        self.resize(920, 760)
        self.setStyleSheet(make_style(self.config.get("manager_theme", "Blackout Blue")))
        self.build_ui()
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.tick)
        self.timer.start(33)

        self.exit_shortcut = QShortcut(QKeySequence("Ctrl+Q"), self)
        self.exit_shortcut.activated.connect(self.force_exit)

    def lang(self):
        return self.config.get("language", "en")

    def build_ui(self):
        self.vehicle_status_labels = []
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(16,16,16,16)
        root.setSpacing(12)
        header = QFrame(); header.setObjectName("Header")
        hl = QVBoxLayout(header)
        title = QLabel("ONYX"); title.setObjectName("TitleLabel")
        sub = QLabel(tr(self.lang(), "subtitle")); sub.setObjectName("SubtitleLabel")
        hl.addWidget(title); hl.addWidget(sub); hl.addWidget(self.make_vehicle_status_label())
        root.addWidget(header)
        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)
        self.tabs.addTab(self.build_general_tab(), tr(self.lang(), "general"))
        self.tabs.addTab(self.build_tiles_tab(), tr(self.lang(), "tiles"))
        self.peak_tab = PeakTab(self)
        self.tabs.addTab(self.peak_tab, tr(self.lang(), "peak"))
        self.prototype_tab = PrototypeLabTab(self)
        self.tabs.addTab(self.prototype_tab, tr(self.lang(), "prototype"))
        self.tabs.addTab(self.build_design_tab(), tr(self.lang(), "design"))
        self.tabs.addTab(self.build_units_tab(), tr(self.lang(), "units"))
        self.tabs.addTab(self.build_stability_tab(), tr(self.lang(), "stability"))
        self.tabs.addTab(self.build_language_tab(), tr(self.lang(), "language"))
        self.tabs.addTab(self.build_hotkeys_tab(), tr(self.lang(), "hotkeys"))
        row = QHBoxLayout()
        root.addLayout(row)
        self.btn_save = QPushButton(tr(self.lang(), "save")); self.btn_save.clicked.connect(self.save_from_forms); row.addWidget(self.btn_save)
        self.btn_overlay = QPushButton(tr(self.lang(), "show_overlay")); self.btn_overlay.clicked.connect(self.show_overlay); row.addWidget(self.btn_overlay)
        self.btn_hide = QPushButton(tr(self.lang(), "hide_overlay")); self.btn_hide.clicked.connect(lambda: self.overlay.hide() if self.overlay else None); row.addWidget(self.btn_hide)
        self.btn_reset = QPushButton(tr(self.lang(), "reset")); self.btn_reset.clicked.connect(self.reset_all); row.addWidget(self.btn_reset)
        self.btn_exit = QPushButton(tr(self.lang(), "exit_app")); self.btn_exit.clicked.connect(self.force_exit); row.addWidget(self.btn_exit)

    def populate_overlay_monitor_select(self):
        try:
            self.overlay_monitor.clear()
            self.overlay_monitor.addItem("Primary Monitor", -1)
            screens = QApplication.screens()
            current = int(self.config.get("overlay_monitor_index", -1))
            for i, screen in enumerate(screens):
                self.overlay_monitor.addItem(overlay_screen_label(i, screen), i)
            idx = self.overlay_monitor.findData(current)
            if idx < 0:
                idx = 0
            self.overlay_monitor.setCurrentIndex(idx)
        except Exception as exc:
            log_error("ManagerWindow.populate_overlay_monitor_select", exc)
            try:
                self.overlay_monitor.addItem("Primary Monitor", -1)
            except Exception:
                pass

    def make_vehicle_status_label(self):
        lab = QLabel(vehicle_summary_text(getattr(self, "latest", None)))
        lab.setWordWrap(True)
        lab.setStyleSheet("font-weight:800; color:#00d9ff; background:rgba(2,8,14,120); border:1px solid rgba(0,217,255,80); border-radius:9px; padding:7px;")
        if not hasattr(self, "vehicle_status_labels"):
            self.vehicle_status_labels = []
        self.vehicle_status_labels.append(lab)
        return lab

    def update_vehicle_status_labels(self):
        try:
            text = vehicle_summary_text(getattr(self, "latest", None))
            for lab in getattr(self, "vehicle_status_labels", []):
                try:
                    lab.setText(text)
                except Exception:
                    pass
        except Exception as exc:
            log_error("ManagerWindow.update_vehicle_status_labels", exc)

    def copy_current_vehicle_entry(self):
        try:
            txt = unknown_vehicle_clipboard_text(getattr(self, "latest", None))
            QApplication.clipboard().setText(txt)
            QMessageBox.information(self, "Vehicle Database", "Current vehicle entry copied to clipboard.")
        except Exception as exc:
            log_error("ManagerWindow.copy_current_vehicle_entry", exc)
            QMessageBox.warning(self, "Vehicle Database", "Could not copy vehicle entry.")

    def open_vehicle_database_folder(self):
        try:
            folder = Path.cwd().resolve()
            if sys.platform.startswith("win"):
                os.startfile(str(folder))
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(folder)])
            else:
                subprocess.Popen(["xdg-open", str(folder)])
        except Exception as exc:
            log_error("ManagerWindow.open_vehicle_database_folder", exc)
            QMessageBox.information(self, "Vehicle Database", f"Database folder:\n{Path.cwd().resolve()}")

    def reload_vehicle_database(self):
        try:
            load_vehicle_database(force_reload=True)
            self.update_vehicle_status_labels()
            total, generated, custom = vehicle_database_stats()
            QMessageBox.information(self, "Vehicle Database", f"Vehicle database reloaded.\n\nTotal IDs: {total}\nGenerated: {generated}\nCustom/override: {custom}")
        except Exception as exc:
            log_error("ManagerWindow.reload_vehicle_database", exc)
            QMessageBox.warning(self, "Vehicle Database", "Could not reload vehicle database.")

    def build_general_tab(self):
        w = QWidget(); outer = QVBoxLayout(w)
        box = QGroupBox(tr(self.lang(), "system")); form = QFormLayout(box); outer.addWidget(box)
        self.udp_port = QSpinBox(); self.udp_port.setRange(1,65535); self.udp_port.setValue(int(self.config["udp_port"]))
        form.addRow(tr(self.lang(),"udp_port")+":", self.udp_port)
        self.opacity = QDoubleSpinBox(); self.opacity.setRange(0.25,1.0); self.opacity.setSingleStep(0.05); self.opacity.setValue(float(self.config["opacity"]))
        form.addRow(tr(self.lang(),"opacity")+":", self.opacity)
        self.scale = QDoubleSpinBox(); self.scale.setRange(0.6,1.8); self.scale.setSingleStep(0.05); self.scale.setValue(float(self.config["scale"]))
        form.addRow(tr(self.lang(),"scale")+":", self.scale)

        self.overlay_monitor = QComboBox()
        self.populate_overlay_monitor_select()
        form.addRow("Overlay Monitor:", self.overlay_monitor)

        form.addRow("Current Vehicle:", self.make_vehicle_status_label())
        self.vehicle_badge_enabled = QCheckBox("show fixed vehicle badge")
        self.vehicle_badge_enabled.setChecked(bool(self.config.get("vehicle_badge_enabled", True)))
        form.addRow("Vehicle Overlay:", self.vehicle_badge_enabled)
        self.vehicle_badge_position = QComboBox()
        self.vehicle_badge_position.addItems(["Top Right", "Bottom Right", "Top Left", "Bottom Left"])
        idx = self.vehicle_badge_position.findText(str(self.config.get("vehicle_badge_position", "Top Right")))
        if idx >= 0:
            self.vehicle_badge_position.setCurrentIndex(idx)
        form.addRow("Vehicle Badge Position:", self.vehicle_badge_position)

        total_db, generated_db, custom_db = vehicle_database_stats()
        self.vehicle_database_info = QLabel(f"{total_db} IDs loaded · generated {generated_db} · custom {custom_db}")
        self.vehicle_database_info.setWordWrap(True)
        form.addRow("Vehicle Database:", self.vehicle_database_info)
        vehicle_db_row = QHBoxLayout()
        btn_copy_vehicle = QPushButton("Copy current/unknown vehicle entry")
        btn_copy_vehicle.clicked.connect(self.copy_current_vehicle_entry)
        vehicle_db_row.addWidget(btn_copy_vehicle)
        btn_open_vehicle_db = QPushButton("Open DB folder")
        btn_open_vehicle_db.clicked.connect(self.open_vehicle_database_folder)
        vehicle_db_row.addWidget(btn_open_vehicle_db)
        btn_reload_vehicle_db = QPushButton("Reload DB")
        btn_reload_vehicle_db.clicked.connect(self.reload_vehicle_database)
        vehicle_db_row.addWidget(btn_reload_vehicle_db)
        form.addRow(vehicle_db_row)

        self.bg_alpha = QSpinBox(); self.bg_alpha.setRange(0,255); self.bg_alpha.setValue(int(self.config["background_alpha"]))
        form.addRow(tr(self.lang(),"tile_bg")+":", self.bg_alpha)
        self.edit_mode = QCheckBox("enabled"); self.edit_mode.setChecked(bool(self.config["edit_mode"]))
        form.addRow(tr(self.lang(),"edit_mode")+":", self.edit_mode)
        self.click = QCheckBox("enabled"); self.click.setChecked(bool(self.config["click_through"]))
        form.addRow(tr(self.lang(),"clickthrough")+":", self.click)
        info = QLabel(tr(self.lang(),"one_exe_note")); info.setWordWrap(True); form.addRow("Info:", info)
        outer.addStretch(1); return w

    def build_tiles_tab(self):
        w = QWidget(); layout = QVBoxLayout(w)
        layout.addWidget(self.make_vehicle_status_label())
        box_sel = QGroupBox(tr(self.lang(),"select_tile")); row = QHBoxLayout(box_sel)
        self.card_select = QComboBox()
        for key,label in CARD_LABELS.items(): self.card_select.addItem(label, key)
        self.card_select.currentIndexChanged.connect(self.load_card_to_form)
        row.addWidget(QLabel("Element:")); row.addWidget(self.card_select); layout.addWidget(box_sel)
        box = QGroupBox(tr(self.lang(),"position")); form = QFormLayout(box); layout.addWidget(box)
        self.card_visible = QCheckBox(tr(self.lang(),"visible"))
        self.card_x = QSpinBox(); self.card_x.setRange(-5000,5000)
        self.card_y = QSpinBox(); self.card_y.setRange(-5000,5000)
        self.card_w = QSpinBox(); self.card_w.setRange(80,2000)
        self.card_h = QSpinBox(); self.card_h.setRange(50,1000)
        self.card_label = QLineEdit()
        self.card_label_visible = QCheckBox("Show label")
        self.card_label_auto = QCheckBox("Auto unit label")
        self.card_color_btn = QPushButton(tr(self.lang(),"choose_color")); self.card_color_btn.clicked.connect(self.choose_color)
        form.addRow(tr(self.lang(),"visible")+":", self.card_visible)
        form.addRow("X:", self.card_x); form.addRow("Y:", self.card_y)
        form.addRow(tr(self.lang(),"width")+":", self.card_w); form.addRow(tr(self.lang(),"height")+":", self.card_h)
        form.addRow("Label visible:", self.card_label_visible)
        form.addRow("Auto label:", self.card_label_auto)
        form.addRow(tr(self.lang(),"label")+":", self.card_label); form.addRow(tr(self.lang(),"color")+":", self.card_color_btn)
        btn = QPushButton(tr(self.lang(),"apply_tile")); btn.clicked.connect(self.apply_current_card); layout.addWidget(btn)
        layout.addStretch(1); self.load_card_to_form(); return w

    def build_design_tab(self):
        w = QWidget(); outer = QVBoxLayout(w)
        outer.addWidget(self.make_vehicle_status_label())
        box = QGroupBox("Manager Design / Theme"); form = QFormLayout(box); outer.addWidget(box)
        self.theme_select = QComboBox()
        for name in THEMES: self.theme_select.addItem(name)
        idx = self.theme_select.findText(self.config.get("manager_theme","Blackout Blue"))
        if idx >= 0: self.theme_select.setCurrentIndex(idx)
        self.theme_select.currentTextChanged.connect(self.apply_theme_preview)
        form.addRow(tr(self.lang(),"theme")+":", self.theme_select)
        outer.addStretch(1); return w


    def build_units_tab(self):
        w = QWidget()
        outer = QVBoxLayout(w)
        outer.addWidget(self.make_vehicle_status_label())

        box = QGroupBox(tr(self.lang(), "unit_system"))
        form = QFormLayout(box)
        outer.addWidget(box)

        self.unit_system_select = QComboBox()
        for val, label_key in [("Metric", "metric"), ("Imperial", "imperial"), ("Custom", "custom")]:
            self.unit_system_select.addItem(tr(self.lang(), label_key), val)
        idx = self.unit_system_select.findData(self.config.get("unit_system", "Metric"))
        if idx >= 0:
            self.unit_system_select.setCurrentIndex(idx)
        self.unit_system_select.currentIndexChanged.connect(self.apply_unit_preset_from_combo)
        form.addRow(tr(self.lang(), "unit_system") + ":", self.unit_system_select)

        self.speed_unit_select = QComboBox()
        for val in ["KMH", "MPH"]:
            self.speed_unit_select.addItem(val, val)
        idx = self.speed_unit_select.findData(self.config.get("speed_unit", "KMH"))
        if idx >= 0:
            self.speed_unit_select.setCurrentIndex(idx)
        form.addRow(tr(self.lang(), "speed_unit") + ":", self.speed_unit_select)

        self.power_unit_select = QComboBox()
        for val in ["PS", "HP", "kW"]:
            self.power_unit_select.addItem(val, val)
        idx = self.power_unit_select.findData(self.config.get("power_unit", "PS"))
        if idx >= 0:
            self.power_unit_select.setCurrentIndex(idx)
        form.addRow(tr(self.lang(), "power_unit") + ":", self.power_unit_select)

        self.boost_unit_select = QComboBox()
        for val in ["bar", "PSI"]:
            self.boost_unit_select.addItem(val, val)
        idx = self.boost_unit_select.findData(self.config.get("boost_unit", "bar"))
        if idx >= 0:
            self.boost_unit_select.setCurrentIndex(idx)
        form.addRow(tr(self.lang(), "boost_unit") + ":", self.boost_unit_select)

        self.gear_label_select = QComboBox()
        for val in ["GEAR", "GANG"]:
            self.gear_label_select.addItem(val, val)
        idx = self.gear_label_select.findData(self.config.get("gear_label", "GEAR"))
        if idx >= 0:
            self.gear_label_select.setCurrentIndex(idx)
        form.addRow(tr(self.lang(), "gear_label") + ":", self.gear_label_select)

        hint = QLabel(tr(self.lang(), "units_hint"))
        hint.setWordWrap(True)
        form.addRow("Info:", hint)

        btn_metric = QPushButton("Metric: KMH / PS / bar")
        btn_metric.clicked.connect(lambda: self.set_unit_preset("Metric"))
        form.addRow(btn_metric)

        btn_imperial = QPushButton("Imperial: MPH / HP / PSI")
        btn_imperial.clicked.connect(lambda: self.set_unit_preset("Imperial"))
        form.addRow(btn_imperial)

        outer.addStretch(1)
        return w

    def apply_unit_preset_from_combo(self):
        if not hasattr(self, "unit_system_select"):
            return
        system = self.unit_system_select.currentData()
        if system in ("Metric", "Imperial"):
            self.set_unit_preset(system, save_now=False)

    def set_unit_preset(self, system, save_now=True):
        if not hasattr(self, "speed_unit_select"):
            return
        if system == "Metric":
            self.unit_system_select.setCurrentIndex(self.unit_system_select.findData("Metric"))
            self.speed_unit_select.setCurrentIndex(self.speed_unit_select.findData("KMH"))
            self.power_unit_select.setCurrentIndex(self.power_unit_select.findData("PS"))
            self.boost_unit_select.setCurrentIndex(self.boost_unit_select.findData("bar"))
            self.gear_label_select.setCurrentIndex(self.gear_label_select.findData("GEAR"))
        elif system == "Imperial":
            self.unit_system_select.setCurrentIndex(self.unit_system_select.findData("Imperial"))
            self.speed_unit_select.setCurrentIndex(self.speed_unit_select.findData("MPH"))
            self.power_unit_select.setCurrentIndex(self.power_unit_select.findData("HP"))
            self.boost_unit_select.setCurrentIndex(self.boost_unit_select.findData("PSI"))
            self.gear_label_select.setCurrentIndex(self.gear_label_select.findData("GEAR"))
        if save_now:
            self.save_from_forms()


    def build_stability_tab(self):
        w = QWidget()
        outer = QVBoxLayout(w)
        outer.addWidget(self.make_vehicle_status_label())
        box = QGroupBox(tr(self.lang(), "stability"))
        form = QFormLayout(box)
        outer.addWidget(box)

        boost_note = QLabel(tr(self.lang(), "boost_fix_note"))
        boost_note.setWordWrap(True)
        form.addRow("BoostFix:", boost_note)

        crash_note = QLabel(tr(self.lang(), "crash_log_hint"))
        crash_note.setWordWrap(True)
        form.addRow(tr(self.lang(), "crash_log") + ":", crash_note)

        perf_note = QLabel(tr(self.lang(), "dyno_perf_note"))
        perf_note.setWordWrap(True)
        form.addRow("Dyno:", perf_note)

        self.crash_log_path_label = QLabel(str(CRASH_LOG_PATH))
        self.crash_log_path_label.setWordWrap(True)
        form.addRow("Path:", self.crash_log_path_label)

        btn_clear_log = QPushButton("Clear crash log")
        btn_clear_log.clicked.connect(self.clear_crash_log)
        form.addRow(btn_clear_log)

        outer.addStretch(1)
        return w

    def clear_crash_log(self):
        try:
            if CRASH_LOG_PATH.exists():
                CRASH_LOG_PATH.unlink()
            QMessageBox.information(self, "OK", "Crash log cleared.")
        except Exception as exc:
            log_error("ManagerWindow.clear_crash_log", exc)
            QMessageBox.warning(self, "Error", "Could not clear crash log.")


    def build_language_tab(self):
        w = QWidget(); outer = QVBoxLayout(w)
        box = QGroupBox("Language / Sprache"); form = QFormLayout(box); outer.addWidget(box)
        self.language_select = QComboBox()
        for code,name in LANGUAGES.items(): self.language_select.addItem(name, code)
        idx = self.language_select.findData(self.lang())
        if idx >= 0: self.language_select.setCurrentIndex(idx)
        form.addRow("Language:", self.language_select)
        hint = QLabel(tr(self.lang(),"lang_hint")); hint.setWordWrap(True); form.addRow("Info:", hint)
        btn = QPushButton("Apply / Anwenden"); btn.clicked.connect(self.apply_language); form.addRow(btn)
        outer.addStretch(1); return w

    def build_hotkeys_tab(self):
        w = QWidget(); outer = QVBoxLayout(w)
        box = QGroupBox("Keyboard shortcuts"); form = QFormLayout(box); outer.addWidget(box)
        hk = self.config.get("hotkeys",{})
        self.hk_toggle_edit = QLineEdit(hk.get("toggle_edit","Ctrl+E"))
        self.hk_toggle_click = QLineEdit(hk.get("toggle_click","Ctrl+T"))
        self.hk_save = QLineEdit(hk.get("save_layout","Ctrl+S"))
        self.hk_reset = QLineEdit(hk.get("reset_layout","Ctrl+R"))
        self.hk_hide = QLineEdit(hk.get("hide_overlay","Esc"))
        form.addRow("Edit mode:", self.hk_toggle_edit); form.addRow("Click-through:", self.hk_toggle_click)
        form.addRow("Save layout:", self.hk_save); form.addRow("Reset layout:", self.hk_reset); form.addRow("Hide overlay:", self.hk_hide)
        outer.addStretch(1); return w

    def current_card_key(self): return self.card_select.currentData()

    def load_card_to_form(self):
        if not hasattr(self, "card_select"): return
        key = self.current_card_key()
        if not key: return
        c = self.config["cards"][key]
        self.card_visible.setChecked(bool(c.get("visible",True)))
        self.card_x.setValue(int(c.get("x",0))); self.card_y.setValue(int(c.get("y",0)))
        self.card_w.setValue(int(c.get("w",230))); self.card_h.setValue(int(c.get("h",88)))
        self.card_label.setText(str(c.get("label", CARD_LABELS.get(key,key))))
        if hasattr(self, "card_label_visible"):
            self.card_label_visible.setChecked(bool(c.get("label_visible", True)))
        if hasattr(self, "card_label_auto"):
            self.card_label_auto.setChecked(bool(c.get("label_auto", True)))
        self.card_color_btn.setStyleSheet(f"background-color:{c.get('color','#fff')}; color:#001018;")

    def reload_forms_from_config(self):
        if hasattr(self, "card_select"):
            self.load_card_to_form()

    def apply_current_card(self):
        if not hasattr(self, "card_select"): return
        key = self.current_card_key()
        c = self.config["cards"][key]
        c["visible"] = self.card_visible.isChecked()
        c["x"] = int(self.card_x.value()); c["y"] = int(self.card_y.value())
        c["w"] = int(self.card_w.value()); c["h"] = int(self.card_h.value())
        c["label"] = self.card_label.text().strip()
        if hasattr(self, "card_label_visible"):
            c["label_visible"] = self.card_label_visible.isChecked()
        if hasattr(self, "card_label_auto"):
            c["label_auto"] = self.card_label_auto.isChecked()

    def choose_color(self):
        key = self.current_card_key()
        col = QColorDialog.getColor(QColor(self.config["cards"][key].get("color","#fff")), self)
        if col.isValid():
            self.config["cards"][key]["color"] = col.name()
            self.card_color_btn.setStyleSheet(f"background-color:{col.name()}; color:#001018;")

    def collect_forms(self):
        self.apply_current_card()
        self.config["udp_port"] = int(self.udp_port.value())
        self.config["opacity"] = float(self.opacity.value())
        self.config["scale"] = float(self.scale.value())
        self.config["background_alpha"] = int(self.bg_alpha.value())
        if hasattr(self, "overlay_monitor"):
            self.config["overlay_monitor_index"] = int(self.overlay_monitor.currentData())
        if hasattr(self, "vehicle_badge_enabled"):
            self.config["vehicle_badge_enabled"] = self.vehicle_badge_enabled.isChecked()
        if hasattr(self, "vehicle_badge_position"):
            self.config["vehicle_badge_position"] = self.vehicle_badge_position.currentText()
        self.config["edit_mode"] = self.edit_mode.isChecked()
        self.config["click_through"] = self.click.isChecked()
        self.config["manager_theme"] = self.theme_select.currentText() if hasattr(self,"theme_select") else self.config.get("manager_theme","Blackout Blue")
        self.config["language"] = self.language_select.currentData() if hasattr(self,"language_select") else self.lang()
        if hasattr(self, "unit_system_select"):
            self.config["unit_system"] = self.unit_system_select.currentData()
            self.config["speed_unit"] = self.speed_unit_select.currentData()
            self.config["power_unit"] = self.power_unit_select.currentData()
            self.config["boost_unit"] = self.boost_unit_select.currentData()
            self.config["gear_label"] = self.gear_label_select.currentData()
            if (self.config["speed_unit"], self.config["power_unit"], self.config["boost_unit"]) == ("KMH", "PS", "bar"):
                self.config["unit_system"] = "Metric"
            elif (self.config["speed_unit"], self.config["power_unit"], self.config["boost_unit"]) == ("MPH", "HP", "PSI"):
                self.config["unit_system"] = "Imperial"
            else:
                self.config["unit_system"] = "Custom"
        self.config["hotkeys"] = {
            "toggle_edit": self.hk_toggle_edit.text().strip(),
            "toggle_click": self.hk_toggle_click.text().strip(),
            "save_layout": self.hk_save.text().strip(),
            "reset_layout": self.hk_reset.text().strip(),
            "hide_overlay": self.hk_hide.text().strip(),
        }

    def save_from_forms(self):
        old_port = self.config["udp_port"]

        # Preserve positions that were changed directly in the overlay before
        # collecting form values. Without this, dragged tiles may revert after restart.
        if self.overlay and hasattr(self.overlay, "cards"):
            for key, card in self.overlay.cards.items():
                if key in self.config.get("cards", {}):
                    self.config["cards"][key]["x"] = int(card.cfg.get("x", self.config["cards"][key].get("x", 0)))
                    self.config["cards"][key]["y"] = int(card.cfg.get("y", self.config["cards"][key].get("y", 0)))
                    self.config["cards"][key]["w"] = int(card.cfg.get("w", self.config["cards"][key].get("w", 230)))
                    self.config["cards"][key]["h"] = int(card.cfg.get("h", self.config["cards"][key].get("h", 88)))

        self.collect_forms()
        save_config(self.config)
        if int(old_port) != int(self.config["udp_port"]):
            self.restart_receiver()
        if self.overlay:
            self.overlay.sync_config()
        QMessageBox.information(self, tr(self.lang(),"saved"), tr(self.lang(),"saved_msg") + "\n\nTip: If the overlay is hidden after saving, press Show Overlay again.")

    def save_config_now(self):
        save_config(self.config)

    def apply_language(self):
        self.collect_forms()
        save_config(self.config)
        QMessageBox.information(self, "OK", "Language saved. Manager will restart.")
        self.rebuild_ui()

    def apply_theme_preview(self, name):
        self.config["manager_theme"] = name
        self.setStyleSheet(make_style(name))

    def rebuild_ui(self):
        self.setWindowTitle(tr(self.lang(), "title"))
        old_overlay_visible = self.overlay.isVisible() if self.overlay else False
        self.takeCentralWidget()
        self.build_ui()
        self.setStyleSheet(make_style(self.config.get("manager_theme","Blackout Blue")))
        if self.overlay:
            self.overlay.sync_config()
            if old_overlay_visible: self.overlay.show()

    def show_overlay(self):
        try:
            # Keep v4.8 behavior: overlay window must cover the whole screen.
            # Tiles are drawn inside this transparent desktop-sized layer.
            # Do NOT resize the overlay to the tile column size, or tiles get clipped.
            if hasattr(self, "save_from_forms"):
                self.save_from_forms()
            elif hasattr(self, "collect_forms"):
                self.collect_forms()
                save_config(self.config)

            if self.overlay is None:
                self.overlay = OverlayWindow(self)

            screen = get_overlay_screen(self.config)
            if screen:
                self.overlay.setGeometry(screen.geometry())
            else:
                self.overlay.setGeometry(0, 0, 1920, 1080)

            self.overlay.sync_config()
            self.overlay.show()
            self.overlay.raise_()
        except Exception as exc:
            log_error("OverlayFullScreenRestore.show_overlay", exc)
            QMessageBox.critical(
                self,
                "Overlay start failed",
                "The overlay could not be started.\n\n"
                "A precise crash log was written to logs/onyx_crash.log.\n\n"
                f"Error: {type(exc).__name__}: {exc}"
            )


    def toggle_overlay_visibility(self):
        try:
            if not self.overlay:
                self.show_overlay()
                return
            if self.overlay.isVisible():
                self.overlay.hide()
            else:
                screen = get_overlay_screen(self.config)
                if screen:
                    self.overlay.setGeometry(screen.geometry())
                else:
                    self.overlay.setGeometry(0, 0, 1920, 1080)
                self.overlay.sync_config()
                self.overlay.show()
                self.overlay.raise_()
        except Exception as exc:
            log_error("ManagerWindow.toggle_overlay_visibility", exc)


    def toggle_recording(self):
        try:
            if hasattr(self, "peak_tab"):
                if self.peak_tab.recording:
                    self.peak_tab.stop()
                else:
                    self.peak_tab.start()
        except Exception as exc:
            log_error("ManagerWindow.toggle_recording", exc)

    def reset_peak_recording(self):
        try:
            if hasattr(self, "peak_tab"):
                self.peak_tab.clear()
        except Exception as exc:
            log_error("ManagerWindow.reset_peak_recording", exc)

    # Compatibility wrappers for overlay hotkeys.
    # These prevent startup crashes if old hotkey names still exist in config.
    def save_layout(self):
        try:
            # Pull any live overlay tile positions back into the real config
            # before collect_forms/save_config.
            if self.overlay and hasattr(self.overlay, "cards"):
                for key, card in self.overlay.cards.items():
                    if key in self.config.get("cards", {}):
                        self.config["cards"][key]["x"] = int(card.cfg.get("x", self.config["cards"][key].get("x", 0)))
                        self.config["cards"][key]["y"] = int(card.cfg.get("y", self.config["cards"][key].get("y", 0)))
                        self.config["cards"][key]["w"] = int(card.cfg.get("w", self.config["cards"][key].get("w", 230)))
                        self.config["cards"][key]["h"] = int(card.cfg.get("h", self.config["cards"][key].get("h", 88)))
            if hasattr(self, "collect_forms"):
                self.collect_forms()
            save_config(self.config)
            if self.overlay:
                self.overlay.sync_config()
        except Exception as exc:
            log_error("ManagerWindow.save_layout", exc)

    def reset_layout(self):
        try:
            if "cards" in self.config and "cards" in DEFAULT_CONFIG:
                self.config["cards"] = json.loads(json.dumps(DEFAULT_CONFIG["cards"]))
            save_config(self.config)
            if self.overlay:
                self.overlay.sync_config()
            if hasattr(self, "rebuild_ui"):
                self.rebuild_ui()
        except Exception as exc:
            log_error("ManagerWindow.reset_layout", exc)

    def toggle_click(self):
        try:
            self.config["click_through"] = not bool(self.config.get("click_through", False))
            save_config(self.config)
            if self.overlay:
                self.overlay.sync_config()
        except Exception as exc:
            log_error("ManagerWindow.toggle_click", exc)

    def reset_overlay_position(self):
        try:
            # Restore tile defaults but keep overlay as a full-screen transparent layer.
            if "cards" in self.config and "cards" in DEFAULT_CONFIG:
                self.config["cards"] = json.loads(json.dumps(DEFAULT_CONFIG["cards"]))
            save_config(self.config)
            if self.overlay:
                screen = get_overlay_screen(self.config)
                if screen:
                    self.overlay.setGeometry(screen.geometry())
                else:
                    self.overlay.setGeometry(0, 0, 1920, 1080)
                self.overlay.sync_config()
                self.overlay.show()
                self.overlay.raise_()
        except Exception as exc:
            log_error("ManagerWindow.reset_overlay_position", exc)


    def toggle_edit(self):
        self.config["edit_mode"] = not self.config.get("edit_mode", True)
        save_config(self.config)
        if self.overlay: self.overlay.sync_config()

    def toggle_click(self):
        self.config["click_through"] = not self.config.get("click_through", False)
        save_config(self.config)
        if self.overlay: self.overlay.sync_config()

    def reset_all(self):
        self.config = deep_copy(DEFAULT_CONFIG)
        save_config(self.config)
        self.restart_receiver()
        self.rebuild_ui()

    def restart_receiver(self):
        self.receiver.stop()
        self.receiver = UdpReceiver(self.config["udp_host"], self.config["udp_port"], self.queue)
        self.receiver.start()

    def tick(self):
        changed = False
        processed = 0
        max_process = 250
        try:
            while processed < max_process:
                self.latest = self.queue.get_nowait()
                processed += 1
                changed = True
                if hasattr(self, "peak_tab"):
                    try:
                        self.peak_tab.add_sample(self.latest)
                    except Exception as exc:
                        log_error("ManagerWindow.tick.peak_tab", exc)
                if hasattr(self, "prototype_tab"):
                    try:
                        self.prototype_tab.add_sample(self.latest)
                    except Exception as exc:
                        log_error("ManagerWindow.tick.prototype_tab", exc)
        except Empty:
            pass
        except Exception as exc:
            log_error("ManagerWindow.tick.queue", exc)

        if changed:
            try:
                cid = effective_car_id(getattr(self, "latest", None))
                if cid is not None and cid not in getattr(self, "logged_unknown_vehicle_ids", set()):
                    if log_unknown_vehicle(self.latest):
                        self.logged_unknown_vehicle_ids.add(cid)
            except Exception as exc:
                log_error("ManagerWindow.tick.vehicle_unknown_logger", exc)
            self.update_vehicle_status_labels()

        if changed and self.overlay:
            try:
                self.overlay.telemetry = self.latest
                self.overlay.update()
            except Exception as exc:
                log_error("ManagerWindow.tick.overlay_update", exc)

    def force_exit(self):
        """
        SafeExit / Killswitch:
        - saves config
        - stops UDP socket
        - hides overlay
        - quits Qt
        - hard-exits the Python process so no frozen window/thread remains
        """
        try:
            save_config(self.config)
        except Exception:
            pass

        try:
            if self.receiver:
                self.receiver.stop()
        except Exception:
            pass

        try:
            if self.overlay:
                self.overlay.hide()
        except Exception:
            pass

        try:
            app = QApplication.instance()
            if app:
                app.quit()
        except Exception:
            pass

        os._exit(0)

    def closeEvent(self, event):
        event.accept()
        self.force_exit()


def main():
    app = QApplication(sys.argv)
    if ICON_PATH.exists():
        app.setWindowIcon(QIcon(str(ICON_PATH)))
    w = ManagerWindow()
    w.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
